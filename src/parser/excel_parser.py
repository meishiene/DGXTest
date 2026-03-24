from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from src.parser.models import (
    CaseRecord,
    DictionaryRecord,
    ObjectRecord,
    PageRecord,
    ParsedObjectRepository,
    ParsedTestSuite,
    StepRecord,
    TestDataRecord,
)


REQUIRED_TEST_SHEETS = [
    "Case_Index",
    "Case_Steps",
    "Test_Data",
    "Run_Config",
    "Dictionary",
]
REQUIRED_OBJECT_SHEETS = ["Page_Index", "Element_Objects"]


def parse_test_workbook(workbook_path: Path) -> ParsedTestSuite:
    workbook = _load_workbook(workbook_path)
    sheet_map = _build_sheet_map(workbook.sheetnames)
    _validate_required_sheets(sheet_map, REQUIRED_TEST_SHEETS, workbook_path)

    cases = [
        CaseRecord(
            case_id=row["case_id"],
            case_name=row["case_name"],
            module=row["module"],
            sub_module=row["sub_module"],
            feature_name=row["feature_name"],
            priority=row["priority"],
            test_level=row["test_level"],
            tags=_split_csv(row["tags"]),
            status=row["status"],
            automation_flag=row["automation_flag"],
            data_set_id=row["data_set_id"],
            preconditions=row["preconditions"],
            expected_result_summary=row["expected_result_summary"],
            owner=row["owner"],
            env_scope=_split_csv(row["env_scope"]),
            browser_scope=_split_csv(row["browser_scope"]),
            can_parallel=row["can_parallel"],
            retry_policy=row.get("retry_policy", ""),
            require_login=row.get("require_login", ""),
            depends_on_case=row.get("depends_on_case", ""),
            raw=row,
        )
        for row in _sheet_to_dicts(workbook[sheet_map["Case_Index"]])
    ]

    steps = [
        StepRecord(
            case_id=row["case_id"],
            step_no=_to_int(row["step_no"]),
            step_type=row["step_type"],
            step_name=row["step_name"],
            action_key=row["action_key"],
            assert_key=row["assert_key"],
            page_name=row["page_name"],
            target=row["target"],
            target_type=row["target_type"],
            value=row["value"],
            value_type=row["value_type"],
            expected=row["expected"],
            expected_type=row["expected_type"],
            match_type=row["match_type"],
            wait=row["wait"],
            timeout=_to_int(row["timeout"], default=10),
            continue_on_fail=row["continue_on_fail"],
            screenshot_on_fail=row["screenshot_on_fail"],
            ai_locator_hint=row["ai_locator_hint"],
            remark=row["remark"],
            raw=row,
        )
        for row in _sheet_to_dicts(workbook[sheet_map["Case_Steps"]])
    ]

    test_data = []
    for row in _sheet_to_dicts(workbook[sheet_map["Test_Data"]]):
        payload = {
            key: value
            for key, value in row.items()
            if key not in {"data_set_id", "data_name", "module", "env", "remark"}
        }
        test_data.append(
            TestDataRecord(
                data_set_id=row["data_set_id"],
                data_name=row["data_name"],
                module=row["module"],
                env=row["env"],
                payload=payload,
                remark=row["remark"],
            )
        )

    run_config_rows = _sheet_to_dicts(workbook[sheet_map["Run_Config"]])
    dictionaries = [
        DictionaryRecord(
            dict_type=row["dict_type"],
            dict_value=row["dict_value"],
            dict_label=row["dict_label"],
            enabled=row["enabled"],
            remark=row["remark"],
        )
        for row in _sheet_to_dicts(workbook[sheet_map["Dictionary"]])
    ]

    return ParsedTestSuite(
        cases=cases,
        steps=steps,
        test_data=test_data,
        run_config_rows=run_config_rows,
        dictionaries=dictionaries,
    )


def parse_object_repository(workbook_path: Path) -> ParsedObjectRepository:
    workbook = _load_workbook(workbook_path)
    sheet_map = _build_sheet_map(workbook.sheetnames)
    _validate_required_sheets(sheet_map, REQUIRED_OBJECT_SHEETS, workbook_path)

    pages = [
        PageRecord(
            page_name=row["page_name"],
            module=row["module"],
            sub_module=row["sub_module"],
            route=row["route"],
            page_title=row["page_title"],
            load_wait_strategy=row["load_wait_strategy"],
            load_timeout_sec=_to_int(row["load_timeout_sec"], default=10),
            frontend_hint=row["frontend_hint"],
            api_group_hint=row["api_group_hint"],
            status=row["status"],
            remark=row["remark"],
        )
        for row in _sheet_to_dicts(workbook[sheet_map["Page_Index"]])
    ]

    objects = [
        ObjectRecord(
            object_key=row["object_key"],
            object_name=row["object_name"],
            page_name=row["page_name"],
            module=row["module"],
            sub_module=row["sub_module"],
            object_type=row["object_type"],
            locator_primary_type=row["locator_primary_type"],
            locator_primary_value=row["locator_primary_value"],
            locator_backup_1_type=row["locator_backup_1_type"],
            locator_backup_1_value=row["locator_backup_1_value"],
            locator_backup_2_type=row["locator_backup_2_type"],
            locator_backup_2_value=row["locator_backup_2_value"],
            default_wait=row["default_wait"],
            default_timeout_sec=_to_int(row["default_timeout_sec"], default=10),
            ai_component_hint=row["ai_component_hint"],
            frontend_file_hint=row["frontend_file_hint"],
            api_hint=row["api_hint"],
            enabled=row["enabled"],
            remark=row["remark"],
        )
        for row in _sheet_to_dicts(workbook[sheet_map["Element_Objects"]])
    ]

    return ParsedObjectRepository(pages=pages, objects=objects)


def _load_workbook(workbook_path: Path):
    if not workbook_path.exists():
        raise FileNotFoundError(f"Excel 文件不存在: {workbook_path}")
    return load_workbook(workbook_path)


def _build_sheet_map(sheet_names: list[str]) -> dict[str, str]:
    sheet_map: dict[str, str] = {}
    for name in sheet_names:
        normalized = _normalize_sheet_name(name)
        if normalized and normalized not in sheet_map:
            sheet_map[normalized] = name
    return sheet_map


def _validate_required_sheets(sheet_map: dict[str, str], required: list[str], workbook_path: Path) -> None:
    missing = [name for name in required if name not in sheet_map]
    if missing:
        raise ValueError(f"Excel 缺少工作表 {missing}: {workbook_path}")


def _sheet_to_dicts(sheet) -> list[dict[str, str]]:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_normalize_header_name(item) for item in rows[0]]
    data_rows: list[dict[str, str]] = []
    for raw_row in rows[1:]:
        if raw_row is None:
            continue
        row = {}
        has_value = False
        for index, header in enumerate(headers):
            if not header:
                continue
            value = raw_row[index] if index < len(raw_row) else ""
            normalized = "" if value is None else str(value).strip()
            row[header] = normalized
            if normalized:
                has_value = True
        if has_value:
            data_rows.append(row)
    return data_rows


def _normalize_sheet_name(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    if " - " in text:
        return text.split(" - ", 1)[0].strip()
    return _normalize_header_name(text)


def _normalize_header_name(value) -> str:
    if value is None:
        return ""
    text = str(value).strip().replace("\r\n", "\n")
    if not text:
        return ""
    first_line = text.split("\n", 1)[0].strip()
    for marker in ("（", "("):
        if marker in first_line:
            first_line = first_line.split(marker, 1)[0].strip()
    return first_line


def _split_csv(value: str) -> list[str]:
    if not value:
        return []
    return [item.strip() for item in value.split(",") if item.strip()]


def _to_int(value: str, default: int = 0) -> int:
    if value == "":
        return default
    return int(float(value))
