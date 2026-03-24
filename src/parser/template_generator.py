from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment


DGX_DEMO_ROUTE = "/dgx/f2b850f5-14a0-4325-b56d-cb1213f63ec2"
PLACEHOLDER_REMARK = "占位断言：后续补充 assert_key、target、expected 和 match_type；未填写的占位断言会被执行器跳过。"

TEST_SHEET_LABELS = {
    "Case_Index": "用例索引",
    "Case_Steps": "用例步骤",
    "Test_Data": "测试数据",
    "Run_Config": "运行配置",
    "Dictionary": "字典配置",
}

OBJECT_SHEET_LABELS = {
    "Page_Index": "页面索引",
    "Element_Objects": "对象库元素",
}

TEST_HEADER_LABELS = {
    "Case_Index": {
        "case_id": "用例ID",
        "case_name": "用例名称",
        "module": "模块",
        "sub_module": "子模块",
        "feature_name": "功能点",
        "case_type": "用例类型",
        "priority": "优先级",
        "test_level": "测试层级",
        "tags": "标签",
        "status": "状态",
        "automation_flag": "自动化标记",
        "precondition_id": "前置条件ID",
        "preconditions": "前置条件",
        "postconditions": "后置条件",
        "data_set_id": "数据集ID",
        "env_scope": "环境范围",
        "browser_scope": "浏览器范围",
        "can_parallel": "是否并行",
        "retry_policy": "重试策略",
        "owner": "负责人",
        "require_login": "是否需要登录",
        "depends_on_case": "依赖用例",
        "expected_result_summary": "期望结果摘要",
        "ai_hint": "AI提示",
        "remark": "备注",
    },
    "Case_Steps": {
        "case_id": "用例ID",
        "step_no": "步骤序号",
        "step_type": "步骤类型",
        "step_name": "步骤名称",
        "action_key": "动作关键字",
        "assert_key": "断言关键字",
        "page_name": "页面名",
        "target": "目标对象",
        "target_type": "目标类型",
        "value": "输入值",
        "value_type": "值类型",
        "expected": "期望结果",
        "expected_type": "期望类型",
        "match_type": "匹配方式",
        "wait": "等待策略",
        "timeout": "超时秒数",
        "continue_on_fail": "失败后继续",
        "screenshot_on_fail": "失败截图",
        "ai_locator_hint": "定位提示",
        "remark": "备注",
    },
    "Test_Data": {
        "data_set_id": "数据集ID",
        "data_name": "数据集名称",
        "module": "模块",
        "env": "环境",
        "project_name": "项目名称",
        "project_result_name": "结果项目名",
        "layer_option": "图层选项",
        "remark": "备注",
    },
    "Run_Config": {
        "config_key": "配置键",
        "config_value": "配置值",
        "remark": "备注",
    },
    "Dictionary": {
        "dict_type": "字典类型",
        "dict_value": "字典值",
        "dict_label": "字典标签",
        "enabled": "是否启用",
        "remark": "备注",
    },
}

OBJECT_HEADER_LABELS = {
    "Page_Index": {
        "page_name": "页面名",
        "module": "模块",
        "sub_module": "子模块",
        "route": "路由",
        "page_title": "页面标题",
        "load_wait_strategy": "加载等待策略",
        "load_timeout_sec": "加载超时秒数",
        "frontend_hint": "前端提示",
        "api_group_hint": "接口分组提示",
        "status": "状态",
        "remark": "备注",
    },
    "Element_Objects": {
        "object_key": "对象Key",
        "object_name": "对象名称",
        "page_name": "页面名",
        "module": "模块",
        "sub_module": "子模块",
        "object_type": "对象类型",
        "locator_primary_type": "主定位器类型",
        "locator_primary_value": "主定位器值",
        "locator_backup_1_type": "备用定位器1类型",
        "locator_backup_1_value": "备用定位器1值",
        "locator_backup_2_type": "备用定位器2类型",
        "locator_backup_2_value": "备用定位器2值",
        "default_wait": "默认等待",
        "default_timeout_sec": "默认超时秒数",
        "ai_component_hint": "AI组件提示",
        "frontend_file_hint": "前端文件提示",
        "api_hint": "接口提示",
        "enabled": "是否启用",
        "remark": "备注",
    },
}

TEST_HEADER_COMMENTS = {
    "Case_Index": {
        "case_id": "用例唯一标识，例如 DGX_001。",
        "case_name": "用例名称，建议写业务可读标题。",
        "module": "一级模块，例如 DGX。",
        "sub_module": "二级模块，例如 Demo。",
        "feature_name": "功能点名称。",
        "case_type": "用例类型，例如 functional。",
        "priority": "优先级，例如 P1 / P2。",
        "test_level": "测试层级，例如 smoke。",
        "tags": "标签，多个值用英文逗号分隔。",
        "status": "用例状态，例如 active。",
        "automation_flag": "是否自动化，Y/N。",
        "precondition_id": "前置条件编号，可为空。",
        "preconditions": "前置条件说明。",
        "postconditions": "后置条件说明，可为空。",
        "data_set_id": "关联测试数据集 ID。",
        "env_scope": "环境范围，多个值用英文逗号分隔。",
        "browser_scope": "浏览器范围，多个值用英文逗号分隔。",
        "can_parallel": "是否允许并行，Y/N。",
        "retry_policy": "失败重试策略标识。",
        "owner": "责任人。",
        "require_login": "是否要求先登录，Y/N。",
        "depends_on_case": "依赖的前置用例 ID，可为空。",
        "expected_result_summary": "期望结果摘要。",
        "ai_hint": "给 AI 的额外提示。",
        "remark": "备注说明。",
    },
    "Case_Steps": {
        "case_id": "所属用例 ID。",
        "step_no": "步骤序号，从 1 开始。",
        "step_type": "步骤类型，常见为 action / assert。",
        "step_name": "步骤名称，建议写成业务动作。",
        "action_key": "动作关键字，例如 click、open_page。",
        "assert_key": "断言关键字，例如 assert_text_contains（断言）。",
        "page_name": "所属页面对象名。",
        "target": "目标对象 key；如果是 open_page，可填写路由。",
        "target_type": "目标类型，例如 element / route。",
        "value": "输入值或动作值。",
        "value_type": "值类型，例如 string。",
        "expected": "期望结果，例如 expected（期望结果）可写断言文本。",
        "expected_type": "期望值类型，例如 string。",
        "match_type": "匹配方式，例如 contains / equals。",
        "wait": "等待策略，例如 visible / dom_ready。",
        "timeout": "超时时间，单位秒。",
        "continue_on_fail": "失败后是否继续，Y/N。",
        "screenshot_on_fail": "失败时是否截图，Y/N。",
        "ai_locator_hint": "给 AI 或定位器的提示。",
        "remark": "步骤备注。",
    },
    "Test_Data": {
        "data_set_id": "数据集唯一 ID。",
        "data_name": "数据集名称。",
        "module": "所属模块。",
        "env": "使用环境。",
        "project_name": "示例业务字段：项目名。",
        "project_result_name": "示例业务字段：结果区域项目名。",
        "layer_option": "示例业务字段：图层选项。",
        "remark": "备注。",
    },
    "Run_Config": {
        "config_key": "配置键。",
        "config_value": "配置值。",
        "remark": "备注。",
    },
    "Dictionary": {
        "dict_type": "字典类型。",
        "dict_value": "字典值。",
        "dict_label": "字典中文含义或展示名。",
        "enabled": "是否启用，Y/N。",
        "remark": "备注。",
    },
}

OBJECT_HEADER_COMMENTS = {
    "Page_Index": {
        "page_name": "页面唯一标识。",
        "module": "一级模块。",
        "sub_module": "二级模块。",
        "route": "页面路由。",
        "page_title": "页面标题。",
        "load_wait_strategy": "页面加载等待策略。",
        "load_timeout_sec": "页面加载超时，单位秒。",
        "frontend_hint": "前端页面提示。",
        "api_group_hint": "相关接口分组提示。",
        "status": "状态，例如 active。",
        "remark": "备注。",
    },
    "Element_Objects": {
        "object_key": "对象唯一 key，步骤中的 target 会引用它。",
        "object_name": "对象中文名称或可读名称。",
        "page_name": "对象所属页面。",
        "module": "一级模块。",
        "sub_module": "二级模块。",
        "object_type": "对象类型，例如 button / checkbox。",
        "locator_primary_type": "主定位器类型，例如 text / xpath / css。",
        "locator_primary_value": "主定位器值。",
        "locator_backup_1_type": "备用定位器 1 类型。",
        "locator_backup_1_value": "备用定位器 1 值。",
        "locator_backup_2_type": "备用定位器 2 类型。",
        "locator_backup_2_value": "备用定位器 2 值。",
        "default_wait": "默认等待策略。",
        "default_timeout_sec": "默认超时，单位秒。",
        "ai_component_hint": "给 AI 的组件提示。",
        "frontend_file_hint": "前端文件提示。",
        "api_hint": "接口提示。",
        "enabled": "是否启用，Y/N。",
        "remark": "备注。",
    },
}

COMMON_GLOSSARY = [
    ("action", "动作步骤", "执行点击、输入、勾选等操作。"),
    ("assert", "断言步骤", "用于校验页面状态或文本是否符合预期。"),
    ("expected", "期望结果", "断言时希望看到的文本、URL 或结果。"),
    ("assert_key", "断言关键字", "例如 assert_text_contains，表示断言方式。"),
    ("action_key", "动作关键字", "例如 click、open_page、check。"),
    ("target", "目标对象", "步骤实际操作或断言的对象 key。"),
    ("object_key", "对象库主键", "对象库里的唯一标识，会被步骤 target 引用。"),
    ("remark", "备注", "补充说明，不参与核心执行逻辑。"),
]


def generate_excel_templates(test_workbook_path: Path, object_repo_path: Path) -> None:
    test_workbook_path.parent.mkdir(parents=True, exist_ok=True)
    object_repo_path.parent.mkdir(parents=True, exist_ok=True)
    _build_test_workbook().save(test_workbook_path)
    _build_object_workbook().save(object_repo_path)


def _build_test_workbook() -> Workbook:
    workbook = Workbook()

    case_sheet = workbook.active
    case_sheet.title = _bilingual_sheet_title("Case_Index", TEST_SHEET_LABELS)
    case_headers = [
        "case_id", "case_name", "module", "sub_module", "feature_name", "case_type", "priority",
        "test_level", "tags", "status", "automation_flag", "precondition_id", "preconditions",
        "postconditions", "data_set_id", "env_scope", "browser_scope", "can_parallel", "retry_policy",
        "owner", "require_login", "depends_on_case", "expected_result_summary", "ai_hint", "remark"
    ]
    case_sheet.append(_bilingual_headers(case_headers, TEST_HEADER_LABELS["Case_Index"]))
    case_sheet.append([
        "DGX_001", "Open DGX demo and reach Set Boundaries", "DGX", "Demo", "Project Entry", "functional", "P1",
        "smoke", "dgx,demo,entry", "active", "Y", "", "Demo URL is reachable", "", "DATA_DGX_DEFAULT_01",
        "demo", "chromium", "Y", "case_retry_1", "tester", "N", "",
        "Set Boundaries panel is reachable", "复用录制好的 DGX demo 冒烟路径", "默认 DGX demo 入口冒烟用例"
    ])
    case_sheet.append([
        "DGX_002", "Configure boundary layers and finish setup", "DGX", "Demo", "Boundary Setup", "functional", "P1",
        "smoke", "dgx,demo,boundary", "active", "Y", "", "Boundary setup page is open", "", "DATA_DGX_DEFAULT_01",
        "demo", "chromium", "Y", "case_retry_1", "tester", "N", "",
        "Boundary setup can be completed and the project result is shown", "复用录制好的 DGX 边界配置流程", "默认 DGX 边界配置回归用例"
    ])
    case_sheet.append([
        "DGX_003", "Verify DGX landing page stable entry state", "DGX", "Demo", "Landing Smoke", "functional", "P1",
        "smoke", "dgx,demo,landing", "active", "Y", "", "Demo URL is reachable", "", "DATA_DGX_DEFAULT_01",
        "demo", "chromium", "Y", "case_retry_1", "tester", "N", "",
        "Landing page keeps the expected route and entry controls visible", "用于检查 DGX 落地页入口状态是否稳定", "默认 DGX 首页稳定性检查"
    ])

    step_sheet = workbook.create_sheet(_bilingual_sheet_title("Case_Steps", TEST_SHEET_LABELS))
    step_headers = [
        "case_id", "step_no", "step_type", "step_name", "action_key", "assert_key", "page_name", "target",
        "target_type", "value", "value_type", "expected", "expected_type", "match_type", "wait", "timeout",
        "continue_on_fail", "screenshot_on_fail", "ai_locator_hint", "remark"
    ]
    step_sheet.append(_bilingual_headers(step_headers, TEST_HEADER_LABELS["Case_Steps"]))
    step_sheet.append(["DGX_001", 1, "action", "Open DGX demo landing page", "open_page", "", "DGXLandingPage", DGX_DEMO_ROUTE, "route", "", "", "", "", "", "dom_ready", 10, "N", "Y", "", ""])
    step_sheet.append(["DGX_001", 2, "assert", "Assert DGX demo route", "", "assert_url_equals", "DGXLandingPage", "", "route", "", "", "https://dgx.xlook.ai/dgx/f2b850f5-14a0-4325-b56d-cb1213f63ec2", "string", "equals", "", 10, "N", "Y", "", "确认已经进入预期的 DGX demo 路由"])
    step_sheet.append(["DGX_001", 3, "action", "Open project card entry button", "click", "", "DGXLandingPage", "project_card_entry_button", "element", "", "", "", "", "", "visible", 10, "N", "Y", "Faya 项目入口按钮", ""])
    step_sheet.append(["DGX_001", 4, "action", "Enter project workspace", "click", "", "DGXLandingPage", "project_go_to_button", "element", "", "", "", "", "", "visible", 10, "N", "Y", "进入项目工作区", ""])
    step_sheet.append(["DGX_001", 5, "assert", "Assert project workspace opened", "", "assert_element_visible", "DGXLandingPage", "set_boundaries_button", "element", "", "", "", "", "", "visible", 10, "N", "Y", "边界设置入口", "点击后应能看到 Set Boundaries 工作流入口按钮"])
    step_sheet.append(["DGX_001", 6, "action", "Open Set Boundaries panel", "click", "", "DGXBoundaryPage", "set_boundaries_button", "element", "", "", "", "", "", "visible", 10, "N", "Y", "Set Boundaries 按钮", ""])
    step_sheet.append(["DGX_001", 7, "assert", "Assert Set Boundaries panel opened", "", "assert_element_visible", "DGXBoundaryPage", "upload_plot_button", "element", "", "", "", "", "", "visible", 10, "N", "Y", "UploadPlot CAD File 按钮", "上传 CAD 文件入口应可见"])
    step_sheet.append(["DGX_001", 8, "action", "Wait upload entry is visible", "wait_element", "", "DGXBoundaryPage", "upload_plot_button", "element", "", "", "", "", "", "visible", 10, "N", "Y", "UploadPlot CAD File 按钮", "等待上传入口稳定可见，避免后续点击抖动"])

    step_sheet.append(["DGX_002", 1, "action", "Open DGX demo landing page", "open_page", "", "DGXLandingPage", DGX_DEMO_ROUTE, "route", "", "", "", "", "", "dom_ready", 10, "N", "Y", "", ""])
    step_sheet.append(["DGX_002", 2, "assert", "Assert DGX demo route", "", "assert_url_equals", "DGXLandingPage", "", "route", "", "", "https://dgx.xlook.ai/dgx/f2b850f5-14a0-4325-b56d-cb1213f63ec2", "string", "equals", "", 10, "N", "Y", "", "确认已经进入预期的 DGX demo 路由"])
    step_sheet.append(["DGX_002", 3, "action", "Open project card entry button", "click", "", "DGXLandingPage", "project_card_entry_button", "element", "", "", "", "", "", "visible", 10, "N", "Y", "Faya 项目入口按钮", ""])
    step_sheet.append(["DGX_002", 4, "action", "Enter project workspace", "click", "", "DGXLandingPage", "project_go_to_button", "element", "", "", "", "", "", "visible", 10, "N", "Y", "进入项目工作区", ""])
    step_sheet.append(["DGX_002", 5, "action", "Open Set Boundaries panel", "click", "", "DGXBoundaryPage", "set_boundaries_button", "element", "", "", "", "", "", "visible", 10, "N", "Y", "Set Boundaries 按钮", ""])
    step_sheet.append(["DGX_002", 6, "action", "Open plot upload dialog", "click", "", "DGXBoundaryPage", "upload_plot_button", "element", "", "", "", "", "", "visible", 10, "N", "Y", "UploadPlot CAD File 按钮", ""])
    step_sheet.append(["DGX_002", 7, "assert", "Assert upload dialog opened", "", "assert_element_visible", "DGXBoundaryPage", "layer_select_button_1", "element", "", "", "", "", "", "visible", 10, "N", "Y", "第一层图层选择器", "上传弹窗打开后应出现图层选择下拉框"])
    step_sheet.append(["DGX_002", 8, "action", "Open first layer selector", "click", "", "DGXBoundaryPage", "layer_select_button_1", "element", "", "", "", "", "", "visible", 10, "N", "Y", "第一层图层选择器", ""])
    step_sheet.append(["DGX_002", 9, "action", "Choose Not Required for first layer", "click", "", "DGXBoundaryPage", "not_required_option", "element", "", "", "", "", "", "visible", 10, "N", "Y", "Not Required 选项", ""])
    step_sheet.append(["DGX_002", 10, "action", "Open second layer selector", "click", "", "DGXBoundaryPage", "layer_select_button_2", "element", "", "", "", "", "", "visible", 10, "N", "Y", "第二层图层选择器", ""])
    step_sheet.append(["DGX_002", 11, "action", "Choose Not Required for second layer", "click", "", "DGXBoundaryPage", "not_required_option", "element", "", "", "", "", "", "visible", 10, "N", "Y", "Not Required 选项", ""])
    step_sheet.append(["DGX_002", 12, "action", "Close upload dialog", "click", "", "DGXBoundaryPage", "cancel_button", "element", "", "", "", "", "", "visible", 10, "N", "Y", "取消按钮", ""])
    step_sheet.append(["DGX_002", 13, "assert", "Assert upload dialog closed", "", "assert_element_hidden", "DGXBoundaryPage", "layer_select_button_1", "element", "", "", "", "", "", "hidden", 10, "N", "Y", "第一层图层选择器", "关闭弹窗后图层选择器应不可见"])
    step_sheet.append(["DGX_002", 14, "action", "Go to next setup step", "click", "", "DGXBoundaryPage", "next_button", "element", "", "", "", "", "", "visible", 10, "N", "Y", "下一步按钮", ""])
    step_sheet.append(["DGX_002", 15, "assert", "Assert layer visibility step opened", "", "assert_element_visible", "DGXBoundaryPage", "plot_area_assignment_active_label", "element", "", "", "", "", "", "visible", 10, "N", "Y", "Plot Area Assignment 激活标签", "点击 Next 后左侧步骤条应高亮到 Plot Area Assignment"])
    step_sheet.append(["DGX_002", 16, "action", "Go to layer visibility step", "click", "", "DGXBoundaryPage", "next_button", "element", "", "", "", "", "", "visible", 10, "N", "Y", "下一步按钮", ""])
    step_sheet.append(["DGX_002", 17, "action", "Hide all layers", "uncheck", "", "DGXBoundaryPage", "show_hide_all_checkbox", "element", "", "", "", "", "", "visible", 10, "N", "Y", "显示/隐藏全部复选框", ""])
    step_sheet.append(["DGX_002", 18, "action", "Enable first layer", "check", "", "DGXBoundaryPage", "layer_checkbox_1", "element", "", "", "", "", "", "visible", 10, "N", "Y", "第一层复选框", ""])
    step_sheet.append(["DGX_002", 19, "action", "Enable second layer", "check", "", "DGXBoundaryPage", "layer_checkbox_2", "element", "", "", "", "", "", "visible", 10, "N", "Y", "第二层复选框", ""])
    step_sheet.append(["DGX_002", 20, "action", "Show all layers", "check", "", "DGXBoundaryPage", "show_hide_all_checkbox", "element", "", "", "", "", "", "visible", 10, "N", "Y", "显示/隐藏全部复选框", ""])
    step_sheet.append(["DGX_002", 21, "action", "Finish boundary setup", "click", "", "DGXBoundaryPage", "done_button", "element", "", "", "", "", "", "visible", 10, "N", "Y", "完成按钮", ""])
    step_sheet.append(["DGX_002", 22, "action", "Wait design result area updated", "wait_element", "", "DGXBoundaryPage", "project_result_cell", "element", "", "", "", "", "", "visible", 10, "N", "Y", "结果项目名称单元格", "等待结果区域刷新，避免立即断言拿到旧值"])
    step_sheet.append(["DGX_002", 23, "assert", "Verify result project cell", "", "assert_text_equals", "DGXBoundaryPage", "project_result_cell", "element", "", "", "${project_result_name}", "string", "equals", "visible", 10, "N", "Y", "结果项目名称单元格", "点击 Done 后结果区域应展示测试数据中的项目名称"])

    step_sheet.append(["DGX_003", 1, "action", "Open DGX demo landing page", "open_page", "", "DGXLandingPage", DGX_DEMO_ROUTE, "route", "", "", "", "", "", "dom_ready", 10, "N", "Y", "", ""])
    step_sheet.append(["DGX_003", 2, "assert", "Assert DGX demo route", "", "assert_url_equals", "DGXLandingPage", "", "route", "", "", "https://dgx.xlook.ai/dgx/f2b850f5-14a0-4325-b56d-cb1213f63ec2", "string", "equals", "", 10, "N", "Y", "", "确认已经进入预期的 DGX demo 路由"])
    step_sheet.append(["DGX_003", 3, "action", "Open project card entry button", "click", "", "DGXLandingPage", "project_card_entry_button", "element", "", "", "", "", "", "visible", 10, "N", "Y", "Faya 项目入口按钮", "项目卡片应展示 Go to Configuration 入口按钮"])
    step_sheet.append(["DGX_003", 4, "assert", "Assert go to button visible", "", "assert_element_visible", "DGXLandingPage", "project_go_to_button", "element", "", "", "", "", "", "visible", 10, "N", "Y", "进入项目工作区按钮", "确认首页卡片仍能看到 Go to Configuration 按钮"])

    data_sheet = workbook.create_sheet(_bilingual_sheet_title("Test_Data", TEST_SHEET_LABELS))
    data_headers = ["data_set_id", "data_name", "module", "env", "project_name", "project_result_name", "layer_option", "remark"]
    data_sheet.append(_bilingual_headers(data_headers, TEST_HEADER_LABELS["Test_Data"]))
    data_sheet.append(["DATA_DGX_DEFAULT_01", "DGX 默认 demo 数据", "DGX", "demo", "Faya Al Saadiyat", "Park Hyatt Abu Dhabi Hotel", "Not Required", "用于驱动 DGX demo 场景的默认测试数据"])

    run_sheet = workbook.create_sheet(_bilingual_sheet_title("Run_Config", TEST_SHEET_LABELS))
    run_headers = ["config_key", "config_value", "remark"]
    run_sheet.append(_bilingual_headers(run_headers, TEST_HEADER_LABELS["Run_Config"]))
    run_sheet.append(["target_env", "demo", "目标运行环境"])
    run_sheet.append(["browser", "chromium", "默认浏览器"])
    run_sheet.append(["include_tags", "dgx,demo", "仅执行这些标签的用例"])
    run_sheet.append(["retry_failed_cases", "Y", "失败后允许重跑一次"])

    dictionary_sheet = workbook.create_sheet(_bilingual_sheet_title("Dictionary", TEST_SHEET_LABELS))
    dictionary_headers = ["dict_type", "dict_value", "dict_label", "enabled", "remark"]
    dictionary_sheet.append(_bilingual_headers(dictionary_headers, TEST_HEADER_LABELS["Dictionary"]))
    for row in [
        ("priority", "P1", "高优先级", "Y", ""),
        ("priority", "P2", "中优先级", "Y", ""),
        ("test_level", "smoke", "冒烟", "Y", ""),
        ("status", "active", "启用", "Y", ""),
        ("action_key", "open_page", "打开页面", "Y", ""),
        ("action_key", "click", "点击", "Y", ""),
        ("action_key", "hover", "悬停", "Y", ""),
        ("action_key", "press_key", "按键输入", "Y", ""),
        ("action_key", "clear_and_input", "清空后输入", "Y", ""),
        ("action_key", "input_password", "输入密码", "Y", ""),
        ("action_key", "input_text", "输入文本", "Y", ""),
        ("action_key", "check", "勾选", "Y", ""),
        ("action_key", "uncheck", "取消勾选", "Y", ""),
        ("action_key", "wait_element", "等待元素", "Y", ""),
        ("action_key", "wait_url", "等待 URL", "Y", ""),
        ("action_key", "select_option", "选择选项", "Y", ""),
        ("action_key", "upload_file", "上传文件", "Y", ""),
        ("assert_key", "assert_url_contains", "断言 URL 包含", "Y", ""),
        ("assert_key", "assert_url_equals", "断言 URL 相等", "Y", ""),
        ("assert_key", "assert_element_visible", "断言元素可见", "Y", ""),
        ("assert_key", "assert_element_hidden", "断言元素隐藏", "Y", ""),
        ("assert_key", "assert_element_enabled", "断言元素可用", "Y", ""),
        ("assert_key", "assert_count_equals", "断言数量相等", "Y", ""),
        ("assert_key", "assert_api_called", "断言接口已调用", "Y", ""),
        ("assert_key", "assert_api_status", "断言接口状态码", "Y", ""),
        ("assert_key", "assert_text_contains", "断言文本包含", "Y", ""),
        ("assert_key", "assert_text_equals", "断言文本相等", "Y", ""),
        ("assert_key", "assert_value_equals", "断言取值相等", "Y", ""),
    ]:
        dictionary_sheet.append(list(row))

    _apply_header_comments(workbook, TEST_HEADER_COMMENTS)
    _append_help_sheet(workbook, TEST_SHEET_LABELS)
    return workbook


def _build_object_workbook() -> Workbook:
    workbook = Workbook()

    page_sheet = workbook.active
    page_sheet.title = _bilingual_sheet_title("Page_Index", OBJECT_SHEET_LABELS)
    page_headers = [
        "page_name", "module", "sub_module", "route", "page_title", "load_wait_strategy",
        "load_timeout_sec", "frontend_hint", "api_group_hint", "status", "remark"
    ]
    page_sheet.append(_bilingual_headers(page_headers, OBJECT_HEADER_LABELS["Page_Index"]))
    page_sheet.append([
        "DGXLandingPage", "DGX", "Demo", DGX_DEMO_ROUTE, "DGX Design", "dom_ready", 10,
        "DGX demo 首页", "dgx", "active", "默认 DGX 入口页面"
    ])
    page_sheet.append([
        "DGXBoundaryPage", "DGX", "Demo", DGX_DEMO_ROUTE, "DGX Design", "dom_ready", 10,
        "DGX 边界设置流程", "dgx", "active", "边界设置页状态"
    ])

    object_sheet = workbook.create_sheet(_bilingual_sheet_title("Element_Objects", OBJECT_SHEET_LABELS))
    object_headers = [
        "object_key", "object_name", "page_name", "module", "sub_module", "object_type",
        "locator_primary_type", "locator_primary_value", "locator_backup_1_type", "locator_backup_1_value",
        "locator_backup_2_type", "locator_backup_2_value", "default_wait", "default_timeout_sec",
        "ai_component_hint", "frontend_file_hint", "api_hint", "enabled", "remark"
    ]
    object_sheet.append(_bilingual_headers(object_headers, OBJECT_HEADER_LABELS["Element_Objects"]))
    object_rows = [
        ["project_card_entry_button", "项目卡片入口按钮", "DGXLandingPage", "DGX", "Demo", "button", "xpath", "(//button[normalize-space()='Faya Al Saadiyat'])[2]", "xpath", "(//*[@role='button' and normalize-space()='Faya Al Saadiyat'])[2]", "", "", "visible", 10, "ProjectCard.EntryButton", "", "", "Y", "来自 codegen 录制，第 2 个同名按钮"],
        ["project_go_to_button", "进入配置按钮", "DGXLandingPage", "DGX", "Demo", "button", "text", "Go to Configuration", "partial_text", "Go to Configuration", "", "", "visible", 10, "ProjectCard.GoToButton", "", "", "Y", "打开已选项目配置"],
        ["set_boundaries_button", "Set Boundaries 入口", "DGXBoundaryPage", "DGX", "Demo", "button", "text", "Set Boundaries", "partial_text", "Set Boundaries", "", "", "visible", 10, "BoundaryFlow.Entry", "", "", "Y", "进入边界设置流程"],
        ["upload_plot_button", "上传 CAD 入口", "DGXBoundaryPage", "DGX", "Demo", "button", "text", "UploadPlot CAD File", "partial_text", "UploadPlot CAD File", "", "", "visible", 10, "BoundaryFlow.UploadPlot", "", "", "Y", "打开 CAD 上传弹窗"],
        ["layer_select_button_1", "第一层图层选择器", "DGXBoundaryPage", "DGX", "Demo", "button", "xpath", "(//button[contains(normalize-space(.), '---- Select a Layer ----')])[1]", "partial_text", "---- Select a Layer ----", "", "", "visible", 10, "BoundaryFlow.LayerSelector1", "", "", "Y", "第一个可见的图层选择器"],
        ["layer_select_button_2", "第二层图层选择器", "DGXBoundaryPage", "DGX", "Demo", "button", "xpath", "(//button[contains(normalize-space(.), '---- Select a Layer ----')])[2]", "partial_text", "---- Select a Layer ----", "", "", "visible", 10, "BoundaryFlow.LayerSelector2", "", "", "Y", "第二个可见的图层选择器"],
        ["not_required_option", "Not Required 选项", "DGXBoundaryPage", "DGX", "Demo", "option", "xpath", "(//*[@role='option'][.//span[normalize-space()='Not Required']])[1]", "xpath", "(//span[normalize-space()='Not Required']/ancestor::*[@role='option'])[1]", "", "", "visible", 10, "BoundaryFlow.NotRequired", "", "", "Y", "上传弹窗中共用的选项"],
        ["cancel_button", "取消按钮", "DGXBoundaryPage", "DGX", "Demo", "button", "text", "Cancel", "partial_text", "Cancel", "", "", "visible", 10, "BoundaryFlow.Cancel", "", "", "Y", "关闭上传弹窗"],
        ["next_button", "下一步按钮", "DGXBoundaryPage", "DGX", "Demo", "button", "css", ".lc-footer-next", "xpath", "//main//div[contains(@class, 'lc-footer-next') and normalize-space()='Next']", "", "", "visible", 10, "BoundaryFlow.Next", "", "", "Y", "推进左侧步骤流程"],
        ["plot_area_assignment_active_label", "Plot Area Assignment 激活标签", "DGXBoundaryPage", "DGX", "Demo", "label", "xpath", "//main//div[contains(@class, 'dgx-step-label') and contains(@class, 'active') and normalize-space()='Plot Area Assignment']", "css", ".lc-content .dgx-step-label.active", "", "", "visible", 10, "BoundaryFlow.StepLabel", "", "", "Y", "点击 Next 后应高亮到 Plot Area Assignment"],
        ["show_hide_all_checkbox", "显示/隐藏全部复选框", "DGXBoundaryPage", "DGX", "Demo", "checkbox", "css", ".mkb-all > input[type='checkbox']", "xpath", "//div[contains(@class, 'mkb-all')]/input[@type='checkbox']", "", "", "visible", 10, "BoundaryFlow.ShowHideAll", "", "", "Y", "控制全部图层显隐"],
        ["layer_checkbox_1", "第一层复选框", "DGXBoundaryPage", "DGX", "Demo", "checkbox", "css", ".mkb-list-item > input", "xpath", "(//div[contains(@class, 'mkb-list-item')]//input[@type='checkbox'])[1]", "", "", "visible", 10, "BoundaryFlow.LayerCheckbox1", "", "", "Y", "录制得到的第一个单层复选框"],
        ["layer_checkbox_2", "第二层复选框", "DGXBoundaryPage", "DGX", "Demo", "checkbox", "css", ".mkb-list > div:nth-child(2) > input", "xpath", "(//div[contains(@class, 'mkb-list')]/*[2]//input[@type='checkbox'])[1]", "", "", "visible", 10, "BoundaryFlow.LayerCheckbox2", "", "", "Y", "录制得到的第二个单层复选框"],
        ["done_button", "完成按钮", "DGXBoundaryPage", "DGX", "Demo", "button", "xpath", "//main//div[contains(@class, 'lc-footer-next') and normalize-space()='Done']", "css", ".lc-footer-next", "", "", "visible", 10, "BoundaryFlow.Done", "", "", "Y", "完成边界设置流程"],
        ["project_result_cell", "结果项目名称单元格", "DGXBoundaryPage", "DGX", "Demo", "cell", "text", "Park Hyatt Abu Dhabi Hotel", "partial_text", "Park Hyatt Abu Dhabi Hotel", "", "", "visible", 10, "BoundaryFlow.ResultCell", "", "", "Y", "边界设置完成后的结果单元格"],
    ]
    for row in object_rows:
        object_sheet.append(row)

    _apply_header_comments(workbook, OBJECT_HEADER_COMMENTS)
    _append_help_sheet(workbook, OBJECT_SHEET_LABELS)
    return workbook



def _bilingual_sheet_title(sheet_name: str, labels: dict[str, str]) -> str:
    chinese_label = labels.get(sheet_name, "")
    return f"{sheet_name} - {chinese_label}" if chinese_label else sheet_name

def _bilingual_headers(headers: list[str], labels: dict[str, str]) -> list[str]:
    return [f"{header}\n{labels.get(header, '')}".strip() for header in headers]


def _apply_header_comments(workbook: Workbook, sheet_comments: dict[str, dict[str, str]]) -> None:
    for sheet_name, comments in sheet_comments.items():
        sheet = workbook[_bilingual_sheet_title(sheet_name, TEST_SHEET_LABELS | OBJECT_SHEET_LABELS)]
        for cell in sheet[1]:
            header_key = str(cell.value).split("\n", 1)[0].strip()
            note = comments.get(header_key, "")
            if note:
                cell.comment = Comment(note, "DGXTest")
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        sheet.row_dimensions[1].height = 32
        sheet.freeze_panes = "A2"


def _append_help_sheet(workbook: Workbook, sheet_labels: dict[str, str]) -> None:
    help_sheet = workbook.create_sheet("中文说明")
    help_sheet.append(["类别", "英文名", "中文说明", "补充说明"])
    help_sheet.append(["说明", "parser_compatibility", "兼容性说明", "请勿修改双语表头中的英文第一行关键字；代码解析时以英文关键字为准。中文仅用于阅读辅助。"])
    help_sheet.append([])
    help_sheet.append(["Sheet", "sheet_name", "中文名称", "说明"])
    for sheet_name, label in sheet_labels.items():
        help_sheet.append(["Sheet", sheet_name, label, f"工作表 {sheet_name} = {label}"])
    help_sheet.append([])
    help_sheet.append(["字段", "english_key", "中文含义", "说明"])
    for english_key, chinese_label, detail in COMMON_GLOSSARY:
        help_sheet.append(["字段", english_key, chinese_label, detail])
    help_sheet.freeze_panes = "A2"
