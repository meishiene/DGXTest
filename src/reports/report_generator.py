from __future__ import annotations

from html import escape
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment

from src.core.bootstrap import RunContext
from src.results.models import (
    AppConfig,
    ArtifactRecord,
    BugCaseLink,
    BugRecord,
    CaseResult,
    GenerationOutputs,
    RunInfo,
    StepResult,
)


CASE_RESULTS_SHEET_LABELS = {
    "Case_Results": "\u7528\u4f8b\u7ed3\u679c",
    "Step_Results": "\u6b65\u9aa4\u7ed3\u679c",
    "Case_Stats": "\u7528\u4f8b\u7edf\u8ba1",
}

BUG_SHEET_LABELS = {
    "Bug_Summary": "\u7f3a\u9677\u6c47\u603b",
    "Bug_Case_Map": "\u7f3a\u9677\u7528\u4f8b\u6620\u5c04",
    "Bug_Evidence": "\u7f3a\u9677\u8bc1\u636e",
    "Bug_Stats": "\u7f3a\u9677\u7edf\u8ba1",
}

CASE_RESULTS_HEADER_LABELS = {
    "Case_Results": {
        "run_id": "\u8fd0\u884cID",
        "case_id": "\u7528\u4f8bID",
        "case_name": "\u7528\u4f8b\u540d\u79f0",
        "module": "\u6a21\u5757",
        "sub_module": "\u5b50\u6a21\u5757",
        "priority": "\u4f18\u5148\u7ea7",
        "status": "\u72b6\u6001",
        "failed_step_no": "\u5931\u8d25\u6b65\u9aa4\u5e8f\u53f7",
        "failed_step_name": "\u5931\u8d25\u6b65\u9aa4\u540d\u79f0",
        "failure_category": "\u5931\u8d25\u5206\u7c7b",
        "failure_message": "\u5931\u8d25\u4fe1\u606f",
        "expected_summary": "\u671f\u671b\u6458\u8981",
        "actual_summary": "\u5b9e\u9645\u6458\u8981",
        "bug_id": "\u7f3a\u9677ID",
        "bug_candidate": "\u7f3a\u9677\u5019\u9009",
        "dedup_key": "\u53bb\u91cd\u952e",
        "suspected_layer": "\u7591\u4f3c\u5c42\u7ea7",
        "retry_count": "\u91cd\u8bd5\u6b21\u6570",
        "blocked_by_case_id": "\u963b\u65ad\u4f9d\u8d56\u7528\u4f8bID",
        "artifact_ids": "\u8bc1\u636eID\u5217\u8868",
    },
    "Step_Results": {
        "step_result_id": "\u6b65\u9aa4\u7ed3\u679cID",
        "run_id": "\u8fd0\u884cID",
        "case_id": "\u7528\u4f8bID",
        "step_no": "\u6b65\u9aa4\u5e8f\u53f7",
        "step_name": "\u6b65\u9aa4\u540d\u79f0",
        "step_type": "\u6b65\u9aa4\u7c7b\u578b",
        "page_name": "\u9875\u9762\u540d",
        "action_key": "\u52a8\u4f5c\u5173\u952e\u5b57",
        "assert_key": "\u65ad\u8a00\u5173\u952e\u5b57",
        "target": "\u76ee\u6807\u5bf9\u8c61",
        "expected": "\u671f\u671b\u503c",
        "actual": "\u5b9e\u9645\u503c",
        "status": "\u72b6\u6001",
        "error_type": "\u9519\u8bef\u7c7b\u578b",
        "error_message": "\u9519\u8bef\u4fe1\u606f",
        "page_url": "\u9875\u9762URL",
        "locator_used_type": "\u547d\u4e2d\u5b9a\u4f4d\u7c7b\u578b",
        "locator_used_value": "\u547d\u4e2d\u5b9a\u4f4d\u503c",
        "component_hint": "\u7ec4\u4ef6\u63d0\u793a",
        "frontend_file_hint": "\u524d\u7aef\u6587\u4ef6\u63d0\u793a",
        "api_hint": "API\u63d0\u793a",
        "attempt_no": "\u5c1d\u8bd5\u6b21\u5e8f",
    },
    "Case_Stats": {
        "metric": "\u7edf\u8ba1\u9879",
        "value": "\u7edf\u8ba1\u503c",
    },
}

BUG_HEADER_LABELS = {
    "Bug_Summary": {
        "bug_id": "缺陷ID",
        "title": "标题",
        "module": "模块",
        "sub_module": "子模块",
        "severity": "严重级别",
        "priority": "优先级",
        "status": "状态",
        "root_cause_category": "根因分类",
        "suspected_layer": "疑似层级",
        "affected_case_count": "影响用例数",
        "affected_case_ids": "影响用例ID",
        "primary_case_id": "主用例ID",
        "artifact_count": "证据数量",
        "dedup_key": "去重键",
        "expected_result": "期望结果",
        "actual_result": "实际结果",
        "failure_message": "失败信息",
        "page_url": "页面URL",
        "failed_step_no": "失败步骤序号",
        "failed_step_name": "失败步骤名称",
        "evidence_paths": "证据路径",
        "verification_steps": "验证步骤",
        "change_scope": "修改范围",
        "do_not_change": "禁止改动",
        "contract_reference": "契约参考",
        "template_reference": "模板参考",
        "manifest_path": "Manifest路径",
        "ai_bug_doc_path": "AI缺陷文档路径",
    },
    "Bug_Case_Map": {
        "bug_id": "缺陷ID",
        "run_id": "运行ID",
        "case_id": "用例ID",
        "case_result_id": "用例结果ID",
        "module": "模块",
        "sub_module": "子模块",
        "priority": "优先级",
        "case_status": "用例状态",
        "failure_category": "失败分类",
        "failed_step_no": "失败步骤序号",
        "failed_step_name": "失败步骤名称",
        "is_primary_case": "是否主用例",
    },
    "Bug_Evidence": {
        "bug_id": "缺陷ID",
        "case_id": "用例ID",
        "step_result_id": "步骤结果ID",
        "artifact_type": "证据类型",
        "file_path": "文件路径",
        "description": "描述",
    },
    "Bug_Stats": {
        "metric": "统计项",
        "value": "统计值",
    },
}


def generate_output_artifacts(
    config: AppConfig,
    run_context: RunContext,
    run_info: RunInfo,
    case_results: list[CaseResult],
    step_results: list[StepResult],
    artifacts: list[ArtifactRecord],
    bugs: list[BugRecord],
    bug_case_links: list[BugCaseLink],
) -> GenerationOutputs:
    report_path = run_context.root_dir / "test_report.html"
    case_excel_path = run_context.root_dir / "case_results.xlsx"
    bug_excel_path = run_context.root_dir / "bug_list.xlsx"

    _write_ai_bug_docs(run_context.ai_bugs_dir, bugs, case_results, step_results, artifacts)
    _write_test_report(report_path, run_context, run_info, case_results, artifacts, bugs)
    _write_case_results_excel(case_excel_path, case_results, step_results)
    _write_bug_excel(bug_excel_path, run_context, bugs, bug_case_links, case_results, step_results, artifacts)

    return GenerationOutputs(
        manifest_path=str(run_context.manifest_path),
        test_report_html=str(report_path),
        case_results_excel=str(case_excel_path),
        bug_list_excel=str(bug_excel_path),
        ai_bug_docs_dir=str(run_context.ai_bugs_dir),
        artifacts_dir=str(run_context.artifacts_dir),
    )


def _write_test_report(
    report_path: Path,
    run_context: RunContext,
    run_info: RunInfo,
    case_results: list[CaseResult],
    artifacts: list[ArtifactRecord],
    bugs: list[BugRecord],
) -> None:
    summary = run_info.execution_summary
    contract_path = "docs/references/ai-bug-handoff/output-contract.md"
    template_path = "docs/references/ai-bug-handoff/bug-task-template.md"

    module_rows = _build_status_summary_rows(case_results, "module")
    priority_rows = _build_status_summary_rows(case_results, "priority")
    failure_rows = _build_failure_category_rows(case_results)
    failed_case_rows = _build_failed_case_rows(case_results)
    artifact_rows = _build_artifact_summary_rows(artifacts)
    case_detail_rows = [
        [
            item.case_id,
            item.case_name,
            item.module,
            item.status,
            item.failed_step_name,
            item.failure_message,
            item.bug_id or "-",
        ]
        for item in case_results
    ]
    bug_rows = [
        [
            bug.bug_id,
            bug.title,
            bug.root_cause_category,
            bug.affected_case_count,
            bug.ai_bug_doc_path,
        ]
        for bug in bugs
    ]

    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <title>测试报告 / Test Report - {escape(run_info.run_id)}</title>
  <style>
    body {{ font-family: "Microsoft YaHei", sans-serif; margin: 24px; color: #1f2937; }}
    h1, h2 {{ color: #0f172a; }}
    .cards {{ display: flex; gap: 16px; margin-bottom: 24px; flex-wrap: wrap; }}
    .card {{ background: #f8fafc; border: 1px solid #cbd5e1; border-radius: 10px; padding: 16px; min-width: 160px; }}
    .card strong {{ display: inline-block; margin-bottom: 6px; }}
    table {{ border-collapse: collapse; width: 100%; margin-top: 12px; margin-bottom: 24px; }}
    th, td {{ border: 1px solid #cbd5e1; padding: 8px; text-align: left; vertical-align: top; }}
    th {{ background: #e2e8f0; }}
    .meta-block {{ background: #f8fafc; border: 1px solid #cbd5e1; border-radius: 10px; padding: 16px; margin-bottom: 24px; }}
    .section-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(320px, 1fr)); gap: 20px; }}
  </style>
</head>
<body>
  <h1>测试报告 / Test Report</h1>
  <p><strong>运行ID / Run ID:</strong> {escape(run_info.run_id)}</p>
  <p><strong>项目 / Project:</strong> {escape(run_info.project_name)} | <strong>套件 / Suite:</strong> {escape(run_info.suite_name)} | <strong>浏览器 / Browser:</strong> {escape(run_info.browser)}</p>
  <h2>交接契约 / Handoff Contract</h2>
  <div class="meta-block">
    <p><strong>Manifest / Manifest:</strong> {escape(str(run_context.manifest_path))}</p>
    <p><strong>输出契约 / Output Contract:</strong> {escape(contract_path)}</p>
    <p><strong>交接模板 / Bug Template:</strong> {escape(template_path)}</p>
    <p><strong>AI缺陷目录 / AI Bug Docs:</strong> {escape(str(run_context.ai_bugs_dir))}</p>
  </div>
  <h2>执行总览 / Execution Overview</h2>
  <div class="cards">
    <div class="card"><strong>总用例 / Total Cases</strong><div>{summary.total_cases}</div></div>
    <div class="card"><strong>通过 / Passed</strong><div>{summary.passed_cases}</div></div>
    <div class="card"><strong>失败 / Failed</strong><div>{summary.failed_cases}</div></div>
    <div class="card"><strong>阻断 / Blocked</strong><div>{summary.blocked_cases}</div></div>
    <div class="card"><strong>跳过 / Skipped</strong><div>{summary.skipped_cases}</div></div>
    <div class="card"><strong>未执行 / Not Run</strong><div>{summary.not_run_cases}</div></div>
    <div class="card"><strong>缺陷数 / Bug Count</strong><div>{summary.bug_count}</div></div>
  </div>
  <div class="section-grid">
    <div>
      <h2>模块统计 / Module Summary</h2>
      {_render_html_table(["模块 / module", "总数 / total", "通过 / passed", "失败 / failed", "阻断 / blocked", "跳过 / skipped", "未执行 / not_run"], module_rows)}
    </div>
    <div>
      <h2>优先级统计 / Priority Summary</h2>
      {_render_html_table(["优先级 / priority", "总数 / total", "通过 / passed", "失败 / failed", "阻断 / blocked", "跳过 / skipped", "未执行 / not_run"], priority_rows)}
    </div>
  </div>
  <div class="section-grid">
    <div>
      <h2>失败分类统计 / Failure Category Summary</h2>
      {_render_html_table(["分类 / category", "用例数 / case_count"], failure_rows)}
    </div>
    <div>
      <h2>证据统计 / Artifact Summary</h2>
      {_render_html_table(["证据类型 / artifact_type", "数量 / count"], artifact_rows)}
    </div>
  </div>
  <h2>异常用例摘要 / Failed & Blocked Summary</h2>
  {_render_html_table(["用例ID / case_id", "用例名称 / case_name", "状态 / status", "失败步骤 / failed_step", "失败分类 / failure_category", "缺陷ID / bug_id", "失败信息 / failure_message"], failed_case_rows)}
  <h2>用例明细 / Case Details</h2>
  {_render_html_table(["用例ID / case_id", "用例名称 / case_name", "模块 / module", "状态 / status", "失败步骤 / failed_step", "失败信息 / failure_message", "缺陷ID / bug_id"], case_detail_rows)}
  <h2>缺陷摘要 / Bug Summary</h2>
  {_render_html_table(["缺陷ID / bug_id", "标题 / title", "分类 / category", "影响用例数 / affected_cases", "AI交接文档 / ai_bug_doc_path"], bug_rows)}
</body>
</html>
"""
    report_path.write_text(html, encoding="utf-8")



def _render_html_table(headers: list[str], rows: list[list[object]]) -> str:
    header_html = "".join(f"<th>{escape(str(header))}</th>" for header in headers)
    if not rows:
        body_html = f"<tr><td colspan='{len(headers)}'>无 / None</td></tr>"
    else:
        body_html = "".join(
            "<tr>" + "".join(f"<td>{escape(str(cell))}</td>" for cell in row) + "</tr>"
            for row in rows
        )
    return f"<table><thead><tr>{header_html}</tr></thead><tbody>{body_html}</tbody></table>"



def _build_status_summary_rows(case_results: list[CaseResult], attribute_name: str) -> list[list[object]]:
    buckets: dict[str, dict[str, int]] = {}
    for item in case_results:
        key = getattr(item, attribute_name, "") or "UNKNOWN"
        bucket = buckets.setdefault(
            str(key),
            {"total": 0, "PASSED": 0, "FAILED": 0, "BLOCKED": 0, "SKIPPED": 0, "NOT_RUN": 0},
        )
        bucket["total"] += 1
        bucket[item.status] = bucket.get(item.status, 0) + 1
    return [
        [
            key,
            values["total"],
            values.get("PASSED", 0),
            values.get("FAILED", 0),
            values.get("BLOCKED", 0),
            values.get("SKIPPED", 0),
            values.get("NOT_RUN", 0),
        ]
        for key, values in sorted(buckets.items())
    ]



def _build_failure_category_rows(case_results: list[CaseResult]) -> list[list[object]]:
    counts: dict[str, int] = {}
    for item in case_results:
        category = item.failure_category.strip()
        if not category:
            continue
        counts[category] = counts.get(category, 0) + 1
    return [[key, value] for key, value in sorted(counts.items(), key=lambda item: (-item[1], item[0]))]



def _build_artifact_summary_rows(artifacts: list[ArtifactRecord]) -> list[list[object]]:
    counts: dict[str, int] = {}
    for artifact in artifacts:
        counts[artifact.artifact_type] = counts.get(artifact.artifact_type, 0) + 1
    return [[key, value] for key, value in sorted(counts.items())]



def _build_failed_case_rows(case_results: list[CaseResult]) -> list[list[object]]:
    rows: list[list[object]] = []
    for item in case_results:
        if item.status not in {"FAILED", "BLOCKED"}:
            continue
        rows.append(
            [
                item.case_id,
                item.case_name,
                item.status,
                item.failed_step_name,
                item.failure_category,
                item.bug_id or "-",
                item.failure_message,
            ]
        )
    return rows


def _write_case_results_excel(

    excel_path: Path,
    case_results: list[CaseResult],
    step_results: list[StepResult],
) -> None:
    workbook = Workbook()
    case_sheet = workbook.active
    case_sheet.title = _bilingual_sheet_title("Case_Results", CASE_RESULTS_SHEET_LABELS)
    case_sheet.append(
        _bilingual_headers(
            [
                "run_id",
                "case_id",
                "case_name",
                "module",
                "sub_module",
                "priority",
                "status",
                "failed_step_no",
                "failed_step_name",
                "failure_category",
                "failure_message",
                "expected_summary",
                "actual_summary",
                "bug_id",
                "bug_candidate",
                "dedup_key",
                "suspected_layer",
                "retry_count",
                "blocked_by_case_id",
                "artifact_ids",
            ],
            CASE_RESULTS_HEADER_LABELS["Case_Results"],
        )
    )
    for item in case_results:
        case_sheet.append(
            [
                item.run_id,
                item.case_id,
                item.case_name,
                item.module,
                item.sub_module,
                item.priority,
                item.status,
                item.failed_step_no,
                item.failed_step_name,
                item.failure_category,
                item.failure_message,
                item.expected_summary,
                item.actual_summary,
                item.bug_id,
                "Y" if item.bug_candidate else "N",
                item.dedup_key,
                item.suspected_layer,
                item.retry_count,
                item.blocked_by_case_id,
                ",".join(item.artifact_ids),
            ]
        )

    step_sheet = workbook.create_sheet(_bilingual_sheet_title("Step_Results", CASE_RESULTS_SHEET_LABELS))
    step_sheet.append(
        _bilingual_headers(
            [
                "step_result_id",
                "run_id",
                "case_id",
                "step_no",
                "step_name",
                "step_type",
                "page_name",
                "action_key",
                "assert_key",
                "target",
                "expected",
                "actual",
                "status",
                "error_type",
                "error_message",
                "page_url",
                "locator_used_type",
                "locator_used_value",
                "component_hint",
                "frontend_file_hint",
                "api_hint",
                "attempt_no",
            ],
            CASE_RESULTS_HEADER_LABELS["Step_Results"],
        )
    )
    for item in step_results:
        step_sheet.append(
            [
                item.step_result_id,
                item.run_id,
                item.case_id,
                item.step_no,
                item.step_name,
                item.step_type,
                item.page_name,
                item.action_key,
                item.assert_key,
                item.target,
                item.expected,
                item.actual,
                item.status,
                item.error_type,
                item.error_message,
                item.page_url,
                item.locator_used_type,
                item.locator_used_value,
                item.component_hint,
                item.frontend_file_hint,
                item.api_hint,
                item.attempt_no,
            ]
        )

    stats_sheet = workbook.create_sheet(_bilingual_sheet_title("Case_Stats", CASE_RESULTS_SHEET_LABELS))
    stats_sheet.append(_bilingual_headers(["metric", "value"], CASE_RESULTS_HEADER_LABELS["Case_Stats"]))
    for metric, value in _build_case_result_stat_rows(case_results, step_results):
        stats_sheet.append([metric, value])

    _format_excel_header(case_sheet)
    _format_excel_header(step_sheet)
    _format_excel_header(stats_sheet)
    workbook.save(excel_path)


def _write_bug_excel(
    excel_path: Path,
    run_context: RunContext,
    bugs: list[BugRecord],
    bug_case_links: list[BugCaseLink],
    case_results: list[CaseResult],
    step_results: list[StepResult],
    artifacts: list[ArtifactRecord],
) -> None:
    bug_contract_rows = _build_bug_contract_rows(run_context, bugs, case_results, step_results, artifacts)
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = _bilingual_sheet_title("Bug_Summary", BUG_SHEET_LABELS)
    summary_sheet.append(
        _bilingual_headers(
            [
                "bug_id",
                "title",
                "module",
                "sub_module",
                "severity",
                "priority",
                "status",
                "root_cause_category",
                "suspected_layer",
                "affected_case_count",
                "affected_case_ids",
                "primary_case_id",
                "artifact_count",
                "dedup_key",
                "expected_result",
                "actual_result",
                "failure_message",
                "page_url",
                "failed_step_no",
                "failed_step_name",
                "evidence_paths",
                "verification_steps",
                "change_scope",
                "do_not_change",
                "contract_reference",
                "template_reference",
                "manifest_path",
                "ai_bug_doc_path",
            ],
            BUG_HEADER_LABELS["Bug_Summary"],
        )
    )
    for bug in bugs:
        bug_contract_row = bug_contract_rows[bug.bug_id]
        summary_sheet.append(
            [
                bug.bug_id,
                bug.title,
                bug.module,
                bug.sub_module,
                bug.severity,
                bug.priority,
                bug.status,
                bug.root_cause_category,
                bug.suspected_layer,
                bug.affected_case_count,
                ",".join(bug.affected_case_ids),
                bug.affected_case_ids[0] if bug.affected_case_ids else "",
                len(bug_contract_row["artifacts"]),
                bug.dedup_key,
                bug.expected_result,
                bug.actual_result,
                bug_contract_row["failure_message"],
                bug_contract_row["page_url"],
                bug.failed_step_no,
                bug.failed_step_name,
                bug_contract_row["evidence_paths"],
                bug_contract_row["verification_steps"],
                bug_contract_row["change_scope"],
                bug_contract_row["do_not_change"],
                bug_contract_row["contract_reference"],
                bug_contract_row["template_reference"],
                bug_contract_row["manifest_path"],
                bug.ai_bug_doc_path,
            ]
        )

    case_map_sheet = workbook.create_sheet(_bilingual_sheet_title("Bug_Case_Map", BUG_SHEET_LABELS))
    case_map_sheet.append(
        _bilingual_headers(
            [
                "bug_id",
                "run_id",
                "case_id",
                "case_result_id",
                "module",
                "sub_module",
                "priority",
                "case_status",
                "failure_category",
                "failed_step_no",
                "failed_step_name",
                "is_primary_case",
            ],
            BUG_HEADER_LABELS["Bug_Case_Map"],
        )
    )
    case_results_by_id = {item.case_id: item for item in case_results}
    for link in bug_case_links:
        linked_case = case_results_by_id.get(link.case_id)
        case_map_sheet.append(
            [
                link.bug_id,
                link.run_id,
                link.case_id,
                link.case_result_id,
                linked_case.module if linked_case else "",
                linked_case.sub_module if linked_case else "",
                linked_case.priority if linked_case else "",
                linked_case.status if linked_case else "",
                linked_case.failure_category if linked_case else "",
                link.failed_step_no,
                link.failed_step_name,
                "Y" if link.is_primary_case else "N",
            ]
        )

    evidence_sheet = workbook.create_sheet(_bilingual_sheet_title("Bug_Evidence", BUG_SHEET_LABELS))
    evidence_sheet.append(
        _bilingual_headers(
            [
                "bug_id",
                "case_id",
                "step_result_id",
                "artifact_type",
                "file_path",
                "description",
            ],
            BUG_HEADER_LABELS["Bug_Evidence"],
        )
    )
    for row in bug_contract_rows.values():
        for artifact in row["artifacts"]:
            evidence_sheet.append(
                [
                    row["bug_id"],
                    artifact.case_id,
                    artifact.step_result_id,
                    artifact.artifact_type,
                    artifact.file_path,
                    artifact.description,
                ]
            )

    stats_sheet = workbook.create_sheet(_bilingual_sheet_title("Bug_Stats", BUG_SHEET_LABELS))
    stats_sheet.append(_bilingual_headers(["metric", "value"], BUG_HEADER_LABELS["Bug_Stats"]))
    for metric, value in _build_bug_stat_rows(bugs, bug_case_links, artifacts):
        stats_sheet.append([metric, value])

    _format_excel_header(summary_sheet)
    _format_excel_header(case_map_sheet)
    _format_excel_header(evidence_sheet)
    _format_excel_header(stats_sheet)
    workbook.save(excel_path)



def _build_case_result_stat_rows(
    case_results: list[CaseResult],
    step_results: list[StepResult],
) -> list[tuple[str, int]]:
    rows = [
        ("case_total", len(case_results)),
        ("case_passed", sum(1 for item in case_results if item.status == "PASSED")),
        ("case_failed", sum(1 for item in case_results if item.status == "FAILED")),
        ("case_blocked", sum(1 for item in case_results if item.status == "BLOCKED")),
        ("case_skipped", sum(1 for item in case_results if item.status == "SKIPPED")),
        ("case_not_run", sum(1 for item in case_results if item.status == "NOT_RUN")),
        ("case_bug_candidate", sum(1 for item in case_results if item.bug_candidate)),
        ("step_total", len(step_results)),
        ("step_passed", sum(1 for item in step_results if item.status == "PASSED")),
        ("step_failed", sum(1 for item in step_results if item.status == "FAILED")),
        ("step_skipped", sum(1 for item in step_results if item.status == "SKIPPED")),
        ("max_retry_count", max((item.retry_count for item in case_results), default=0)),
    ]
    return rows


def _build_bug_stat_rows(
    bugs: list[BugRecord],
    bug_case_links: list[BugCaseLink],
    artifacts: list[ArtifactRecord],
) -> list[tuple[str, int]]:
    rows: list[tuple[str, int]] = [
        ("bug_count", len(bugs)),
        ("linked_case_count", len(bug_case_links)),
        ("primary_case_count", sum(1 for item in bug_case_links if item.is_primary_case)),
        ("artifact_count", len(artifacts)),
    ]

    severity_counts: dict[str, int] = {}
    root_cause_counts: dict[str, int] = {}
    layer_counts: dict[str, int] = {}
    for bug in bugs:
        severity_counts[bug.severity] = severity_counts.get(bug.severity, 0) + 1
        root_cause_counts[bug.root_cause_category] = root_cause_counts.get(bug.root_cause_category, 0) + 1
        layer_counts[bug.suspected_layer] = layer_counts.get(bug.suspected_layer, 0) + 1

    rows.extend((f"severity:{key}", value) for key, value in sorted(severity_counts.items()))
    rows.extend((f"root_cause_category:{key}", value) for key, value in sorted(root_cause_counts.items()))
    rows.extend((f"suspected_layer:{key}", value) for key, value in sorted(layer_counts.items()))
    return rows


def _build_bug_contract_rows(
    run_context: RunContext,
    bugs: list[BugRecord],
    case_results: list[CaseResult],
    step_results: list[StepResult],
    artifacts: list[ArtifactRecord],
) -> dict[str, dict[str, object]]:
    case_map = {item.case_id: item for item in case_results}
    steps_by_case: dict[str, list[StepResult]] = {}
    artifacts_by_case: dict[str, list[ArtifactRecord]] = {}

    for step in step_results:
        steps_by_case.setdefault(step.case_id, []).append(step)
    for case_id in steps_by_case:
        steps_by_case[case_id].sort(key=lambda item: item.step_no)
    for artifact in artifacts:
        artifacts_by_case.setdefault(artifact.case_id, []).append(artifact)

    contract_path = "docs/references/ai-bug-handoff/output-contract.md"
    template_path = "docs/references/ai-bug-handoff/bug-task-template.md"
    rows: dict[str, dict[str, object]] = {}
    for bug in bugs:
        primary_case = case_map[bug.affected_case_ids[0]]
        case_steps = steps_by_case.get(primary_case.case_id, [])
        step = next((item for item in case_steps if item.step_no == bug.failed_step_no), None)
        case_artifacts = artifacts_by_case.get(primary_case.case_id, [])
        verification_steps = _build_verification_steps(primary_case)
        rows[bug.bug_id] = {
            "bug_id": bug.bug_id,
            "failure_message": primary_case.failure_message,
            "page_url": step.page_url if step else "",
            "evidence_paths": " | ".join(f"{artifact.artifact_type}: {artifact.file_path}" for artifact in case_artifacts) or "无",
            "verification_steps": " | ".join(verification_steps),
            "change_scope": _build_change_scope(step, bug),
            "do_not_change": _build_do_not_change(bug),
            "contract_reference": contract_path,
            "template_reference": template_path,
            "manifest_path": str(run_context.manifest_path),
            "artifacts": case_artifacts,
        }
    return rows
def _bilingual_sheet_title(sheet_name: str, labels: dict[str, str]) -> str:
    chinese_label = labels.get(sheet_name, "")
    return f"{sheet_name} - {chinese_label}" if chinese_label else sheet_name


def _bilingual_headers(headers: list[str], labels: dict[str, str]) -> list[str]:
    return [f"{header}\n{labels.get(header, '')}".strip() for header in headers]


def _format_excel_header(sheet) -> None:
    for cell in sheet[1]:
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[1].height = 32
    sheet.freeze_panes = "A2"


def _write_ai_bug_docs(
    ai_bugs_dir: Path,
    bugs: list[BugRecord],
    case_results: list[CaseResult],
    step_results: list[StepResult],
    artifacts: list[ArtifactRecord],
) -> None:
    case_map = {item.case_id: item for item in case_results}
    steps_by_case: dict[str, list[StepResult]] = {}
    artifacts_by_case: dict[str, list[ArtifactRecord]] = {}

    for step in step_results:
        steps_by_case.setdefault(step.case_id, []).append(step)
    for case_id in steps_by_case:
        steps_by_case[case_id].sort(key=lambda item: item.step_no)
    for artifact in artifacts:
        artifacts_by_case.setdefault(artifact.case_id, []).append(artifact)

    for bug in bugs:
        primary_case = case_map[bug.affected_case_ids[0]]
        case_steps = steps_by_case.get(primary_case.case_id, [])
        step = next((item for item in case_steps if item.step_no == bug.failed_step_no), None)
        case_artifacts = artifacts_by_case.get(primary_case.case_id, [])

        screenshot_path = _find_artifact_path(case_artifacts, "screenshot")
        dom_snapshot_path = _find_artifact_path(case_artifacts, "dom_snapshot")
        console_log_path = _find_artifact_path(case_artifacts, "console_log")
        network_log_path = _find_artifact_path(case_artifacts, "network_log")
        video_path = _find_artifact_path(case_artifacts, "video")
        evidence_paths = _build_artifact_lines(case_artifacts)
        suspected_files = _build_suspected_files(step)
        change_scope = _build_change_scope(step, bug)
        do_not_change = _build_do_not_change(bug)
        verification_steps = _build_verification_steps(primary_case)
        done_definition = _build_done_definition(primary_case)
        first_check, second_check, third_check = _build_task_handoff(step, bug)
        root_cause_hypothesis = _build_root_cause_hypothesis(primary_case, step, bug)
        issue_judgement = _build_issue_judgement(primary_case)
        repro_steps = _build_repro_steps(case_steps)
        contract_path = "docs/references/ai-bug-handoff/output-contract.md"
        template_path = "docs/references/ai-bug-handoff/bug-task-template.md"

        doc_path = ai_bugs_dir / f"{bug.bug_id}.md"
        bug.ai_bug_doc_path = str(doc_path)
        content = f"""# {bug.bug_id} {bug.title}

## 1. 基本信息
- bug_id: {bug.bug_id}
- title: {bug.title}
- run_id: {bug.run_id}
- case_id: {primary_case.case_id}
- case_name: {primary_case.case_name}
- module: {bug.module}
- sub_module: {bug.sub_module}
- severity: {bug.severity}
- priority: {bug.priority}
- suspected_layer: {bug.suspected_layer}
- contract_reference: {contract_path}
- template_reference: {template_path}

## 2. 问题判定
- 是否为正式 bug: {"是" if primary_case.bug_candidate else "否"}
- 判定依据: {issue_judgement}
- 当前不确定点: {_build_uncertainty(bug)}

## 3. 复现路径
{repro_steps}

## 4. 失败位置
- page_name: {step.page_name if step else ""}
- page_url: {step.page_url if step else ""}
- failed_step_no: {bug.failed_step_no}
- failed_step_name: {bug.failed_step_name}
- target: {step.target if step else ""}
- locator_type: {step.locator_used_type if step else ""}
- locator_value: {step.locator_used_value if step else ""}

## 5. 预期与实际
- expected_result: {bug.expected_result}
- actual_result: {bug.actual_result}
- failure_category: {primary_case.failure_category}
- failure_message: {primary_case.failure_message}

## 6. 证据列表
- screenshot: {screenshot_path}
- dom_snapshot: {dom_snapshot_path}
- console_log: {console_log_path}
- network_log: {network_log_path}
- video: {video_path}
- evidence_paths:
{evidence_paths}

## 7. 定位线索
- suspected_files: {suspected_files}
- component_hint: {step.component_hint if step else ""}
- api_hint: {step.api_hint if step else ""}
- root_cause_hypothesis: {root_cause_hypothesis}

## 8. 任务领取说明
- first_check: {first_check}
- second_check: {second_check}
- third_check: {third_check}

## 9. 修改边界
- change_scope: {change_scope}
- do_not_change: {do_not_change}

## 10. 验证步骤
1. {verification_steps[0]}
2. {verification_steps[1]}
3. {verification_steps[2]}

## 11. 完成定义
- done_definition_1: {done_definition[0]}
- done_definition_2: {done_definition[1]}
- done_definition_3: {done_definition[2]}

## 12. 影响范围和去重信息
- affected_case_count: {bug.affected_case_count}
- affected_case_ids: {", ".join(bug.affected_case_ids)}
- dedup_key: {bug.dedup_key}
"""
        doc_path.write_text(content, encoding="utf-8-sig")


def _build_artifact_lines(artifacts: list[ArtifactRecord]) -> str:
    if not artifacts:
        return "- 无"
    return "\n".join(f"- {artifact.artifact_type}: {artifact.file_path}" for artifact in artifacts)


def _find_artifact_path(artifacts: list[ArtifactRecord], artifact_type: str) -> str:
    for artifact in artifacts:
        if artifact.artifact_type == artifact_type:
            return artifact.file_path
    return ""


def _build_suspected_files(step: StepResult | None) -> str:
    if step is None:
        return ""
    return step.frontend_file_hint or ""


def _build_change_scope(step: StepResult | None, bug: BugRecord) -> str:
    if step is None:
        return f"仅在 {bug.module} 模块当前失败步骤直接相关的逻辑内排查和修改。"
    scope_parts: list[str] = []
    if step.page_name:
        scope_parts.append(f"{step.page_name} 页面逻辑")
    if step.target:
        scope_parts.append(f"{step.target} 对应交互或断言逻辑")
    if step.component_hint:
        scope_parts.append(f"{step.component_hint} 相关组件")
    if step.api_hint:
        scope_parts.append(f"{step.api_hint} 触发链路")
    return "；".join(scope_parts) if scope_parts else f"{bug.module} 模块内与失败步骤直接相关的逻辑"


def _build_do_not_change(bug: BugRecord) -> str:
    if bug.suspected_layer == "frontend":
        return "不要优先修改无证据指向的后端服务、数据库结构或无关业务模块。"
    if bug.suspected_layer == "backend":
        return "不要优先修改无证据指向的前端页面结构或样式逻辑。"
    return "不要优先修改与当前失败步骤无直接证据关联的模块。"


def _build_verification_steps(primary_case: CaseResult) -> list[str]:
    return [
        f"重新执行用例 {primary_case.case_id}，确认原失败步骤恢复通过。",
        f"确认用例 {primary_case.case_id} 的关键预期结果与实际结果一致。",
        "确认原错误信息不再出现，并检查没有引入明显范围外回归。",
    ]


def _build_done_definition(primary_case: CaseResult) -> list[str]:
    return [
        f"用例 {primary_case.case_id} 可以稳定通过。",
        "原关键失败日志或错误现象不再复现。",
        "修改范围内未观察到新的阻断性问题或明显副作用。",
    ]


def _build_task_handoff(step: StepResult | None, bug: BugRecord) -> tuple[str, str, str]:
    if step is None:
        return (
            "先检查失败步骤对应的代码入口。",
            "再检查当前错误信息是否有更直接的代码线索。",
            "最后补充验证步骤确保问题闭环。",
        )
    first_check = f"先检查 {step.frontend_file_hint or step.page_name or bug.module} 中与 {step.target or bug.failed_step_name} 直接相关的实现。"
    second_check = f"再检查 {step.component_hint or step.target or bug.failed_step_name} 的事件绑定、参数传递和状态更新逻辑。"
    third_check = f"最后检查 {step.api_hint or '相关接口链路'} 与页面行为是否一致，确认修改未超出当前边界。"
    return first_check, second_check, third_check


def _build_root_cause_hypothesis(
    primary_case: CaseResult,
    step: StepResult | None,
    bug: BugRecord,
) -> str:
    if step is None:
        return "当前仅能确认存在失败现象，缺少足够步骤上下文。"
    if bug.suspected_layer == "frontend":
        return f"当前更像是前端侧问题，优先检查 {step.target or '目标对象'} 的事件绑定、参数处理和页面状态更新逻辑。"
    return "当前根因仍需进一步结合日志、接口和代码上下文确认。"


def _build_issue_judgement(primary_case: CaseResult) -> str:
    if primary_case.bug_candidate:
        return "当前失败属于产品缺陷候选，且已有失败步骤、失败信息和证据文件支撑。"
    return "当前失败暂不满足正式产品缺陷判定条件。"


def _build_uncertainty(bug: BugRecord) -> str:
    if bug.suspected_layer == "frontend":
        return "当前更偏向前端问题，但仍建议在修复后结合真实链路回归确认接口与数据侧未受影响。"
    return "当前疑似层级仍需结合更多证据进一步确认。"


def _build_repro_steps(case_steps: list[StepResult]) -> str:
    if not case_steps:
        return "1. 当前缺少步骤数据，请回看执行结果。"
    return "\n".join(f"{index}. {step.step_name}" for index, step in enumerate(case_steps, start=1))

