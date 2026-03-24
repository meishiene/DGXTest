from pathlib import Path

from openpyxl import load_workbook

from src.parser.excel_parser import parse_object_repository, parse_test_workbook
from src.parser.template_generator import generate_excel_templates
from src.validator.excel_validator import validate_parsed_workbooks



def test_generate_and_parse_templates(tmp_path: Path) -> None:
    test_workbook = tmp_path / "test_suite_template.xlsx"
    object_workbook = tmp_path / "object_repository_template.xlsx"

    generate_excel_templates(test_workbook, object_workbook)

    suite = parse_test_workbook(test_workbook)
    repository = parse_object_repository(object_workbook)
    validate_parsed_workbooks(suite, repository)

    assert len(suite.cases) == 3
    assert len(suite.steps) == 35
    assert len(suite.test_data) == 1
    assert len(repository.pages) == 2
    assert len(repository.objects) == 15
    assert suite.cases[0].case_id == "DGX_001"
    assert suite.cases[0].retry_policy == "case_retry_1"
    assert suite.cases[0].require_login == "N"
    assert suite.cases[0].depends_on_case == ""
    assert suite.steps[0].step_name == "Open DGX demo landing page"
    route_assert = next(step for step in suite.steps if step.case_id == "DGX_001" and step.step_no == 2)
    result_assert = next(step for step in suite.steps if step.case_id == "DGX_002" and step.step_no == 23)
    landing_case = next(case for case in suite.cases if case.case_id == "DGX_003")
    assert route_assert.assert_key == "assert_url_equals"
    assert route_assert.expected == "https://dgx.xlook.ai/dgx/f2b850f5-14a0-4325-b56d-cb1213f63ec2"
    assert result_assert.assert_key == "assert_text_equals"
    assert result_assert.expected == "${project_result_name}"
    assert landing_case.case_name == "Verify DGX landing page stable entry state"
    assert repository.objects[0].object_key == "project_card_entry_button"
    assert any(item.object_key == "plot_area_assignment_active_label" for item in repository.objects)
    assert any(step.action_key == "wait_element" for step in suite.steps)
    assert any(step.assert_key == "assert_element_hidden" for step in suite.steps)

    supported_dictionary_keys = {(item.dict_type, item.dict_value) for item in suite.dictionaries if item.enabled == "Y"}
    assert ("action_key", "hover") in supported_dictionary_keys
    assert ("action_key", "press_key") in supported_dictionary_keys
    assert ("action_key", "select_option") in supported_dictionary_keys
    assert ("action_key", "upload_file") in supported_dictionary_keys
    assert ("action_key", "wait_url") in supported_dictionary_keys
    assert ("assert_key", "assert_url_equals") in supported_dictionary_keys
    assert ("assert_key", "assert_element_enabled") in supported_dictionary_keys
    assert ("assert_key", "assert_count_equals") in supported_dictionary_keys
    assert ("assert_key", "assert_api_called") in supported_dictionary_keys
    assert ("assert_key", "assert_api_status") in supported_dictionary_keys
    assert ("assert_key", "assert_text_equals") in supported_dictionary_keys

    test_template = load_workbook(test_workbook)
    object_template = load_workbook(object_workbook)

    assert "Case_Index - 用例索引" in test_template.sheetnames
    assert "Case_Steps - 用例步骤" in test_template.sheetnames
    assert "中文说明" in test_template.sheetnames
    assert "Page_Index - 页面索引" in object_template.sheetnames
    assert "Element_Objects - 对象库元素" in object_template.sheetnames
    assert "中文说明" in object_template.sheetnames

    case_steps_sheet = test_template["Case_Steps - 用例步骤"]
    object_sheet = object_template["Element_Objects - 对象库元素"]

    assert case_steps_sheet["F1"].value == "assert_key\n断言关键字"
    assert case_steps_sheet["L1"].value == "expected\n期望结果"
    assert object_sheet["A1"].value == "object_key\n对象Key"
    assert case_steps_sheet["F1"].comment is not None
    assert "断言关键字" in case_steps_sheet["F1"].comment.text
    assert object_sheet["A1"].comment is not None
    assert "对象唯一 key" in object_sheet["A1"].comment.text



def test_validate_parsed_workbooks_rejects_unknown_dependency(tmp_path: Path) -> None:
    test_workbook = tmp_path / "test_suite_template.xlsx"
    object_workbook = tmp_path / "object_repository_template.xlsx"

    generate_excel_templates(test_workbook, object_workbook)

    workbook = load_workbook(test_workbook)
    case_sheet = workbook["Case_Index - 用例索引"]
    depends_on_column = "V"
    case_sheet[f"{depends_on_column}2"] = "CASE_MISSING"
    workbook.save(test_workbook)

    suite = parse_test_workbook(test_workbook)
    repository = parse_object_repository(object_workbook)

    try:
        validate_parsed_workbooks(suite, repository)
        assert False, "expected unknown depends_on_case validation error"
    except ValueError as exc:
        assert "Unknown depends_on_case" in str(exc)
