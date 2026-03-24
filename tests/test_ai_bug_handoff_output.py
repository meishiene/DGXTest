from pathlib import Path
import re

from openpyxl import load_workbook

from src.core.bootstrap import prepare_run_context
from src.reports.report_generator import generate_output_artifacts
from src.results.models import (
    AppConfig,
    ArtifactRecord,
    BugCaseLink,
    BugRecord,
    CaseResult,
    ExecutionSummary,
    RunInfo,
    StepResult,
)

CASE_RESULTS_SHEET = "Case_Results - 用例结果"
STEP_RESULTS_SHEET = "Step_Results - 步骤结果"
CASE_STATS_SHEET = "Case_Stats - 用例统计"
BUG_SUMMARY_SHEET = "Bug_Summary - 缺陷汇总"
BUG_CASE_MAP_SHEET = "Bug_Case_Map - 缺陷用例映射"
BUG_EVIDENCE_SHEET = "Bug_Evidence - 缺陷证据"
BUG_STATS_SHEET = "Bug_Stats - 缺陷统计"


def _sheet_headers(sheet) -> list[str]:
    return [str(cell.value).split("\n", 1)[0] for cell in sheet[1]]


def _sheet_row_as_dict(sheet, row_index: int = 2) -> dict[str, object]:
    headers = _sheet_headers(sheet)
    values = [cell.value for cell in sheet[row_index]]
    return dict(zip(headers, values))


def test_ai_bug_doc_matches_handoff_contract(tmp_path: Path) -> None:
    config = AppConfig(
        project_name="DGX Web Auto Test",
        suite_name="contract_test",
        trigger_by="tester",
        base_url="https://example.com",
        output_root=str(tmp_path / "outputs"),
    )
    run_context = prepare_run_context(config)
    summary = ExecutionSummary(
        total_cases=1,
        passed_cases=0,
        failed_cases=1,
        blocked_cases=0,
        skipped_cases=0,
        not_run_cases=0,
        bug_count=1,
    )
    run_info = RunInfo(
        run_id=run_context.run_id,
        project_name=config.project_name,
        suite_name=config.suite_name,
        trigger_by=config.trigger_by,
        start_time="2026-03-20T14:00:00+08:00",
        end_time="2026-03-20T14:01:00+08:00",
        duration_sec=60,
        environment="test",
        base_url=config.base_url,
        browser=config.browser,
        headless=config.headless,
        execution_summary=summary,
    )
    case_result = CaseResult(
        case_result_id="CR_001",
        run_id=run_context.run_id,
        case_id="LOGIN_002",
        case_name="??????????",
        module="????",
        sub_module="??",
        priority="P1",
        status="FAILED",
        failure_category="JS_ERROR",
        expected_summary="??????????????",
        actual_summary="????????????????",
        failed_step_no=4,
        failed_step_name="??????",
        failure_message="Cannot read properties of undefined (reading 'trim')",
        bug_candidate=True,
        dedup_key="interaction_fail|/login|login_submit_button|click|no_request_sent",
        artifact_ids=["ART_001", "ART_002", "ART_003", "ART_004"],
        suspected_layer="frontend",
        bug_id="BUG-0001",
    )
    step_result = StepResult(
        step_result_id="SR_001",
        run_id=run_context.run_id,
        case_id="LOGIN_002",
        step_no=4,
        step_type="action",
        step_name="??????",
        status="FAILED",
        action_key="click",
        target="login_submit_button",
        page_name="LoginPage",
        error_type="JS_ERROR",
        error_message="Cannot read properties of undefined (reading 'trim')",
        page_url="https://example.com/login",
        locator_used_type="testid",
        locator_used_value="login-submit",
        component_hint="LoginForm.SubmitButton",
        frontend_file_hint="src/components/auth/LoginForm.tsx",
        api_hint="POST /api/auth/login",
    )
    artifacts = [
        ArtifactRecord(
            artifact_id="ART_001",
            run_id=run_context.run_id,
            case_id="LOGIN_002",
            step_result_id="SR_001",
            artifact_type="screenshot",
            file_path=str(run_context.screenshots_dir / "LOGIN_002_failure.png"),
            description="????",
        ),
        ArtifactRecord(
            artifact_id="ART_002",
            run_id=run_context.run_id,
            case_id="LOGIN_002",
            step_result_id="SR_001",
            artifact_type="dom_snapshot",
            file_path=str(run_context.html_snapshots_dir / "LOGIN_002_failure.html"),
            description="DOM ??",
        ),
        ArtifactRecord(
            artifact_id="ART_003",
            run_id=run_context.run_id,
            case_id="LOGIN_002",
            step_result_id="SR_001",
            artifact_type="console_log",
            file_path=str(run_context.console_logs_dir / "LOGIN_002_failure.log"),
            description="Console ??",
        ),
        ArtifactRecord(
            artifact_id="ART_004",
            run_id=run_context.run_id,
            case_id="LOGIN_002",
            step_result_id="SR_001",
            artifact_type="network_log",
            file_path=str(run_context.network_logs_dir / "LOGIN_002_failure.har"),
            description="Network ??",
        ),
    ]
    bug = BugRecord(
        bug_id="BUG-0001",
        run_id=run_context.run_id,
        title="???? - ?????? ????",
        module="????",
        sub_module="??",
        severity="S2",
        priority="P1",
        status="NEW",
        root_cause_category="JS_ERROR",
        suspected_layer="frontend",
        affected_case_ids=["LOGIN_002"],
        affected_case_count=1,
        dedup_key="interaction_fail|/login|login_submit_button|click|no_request_sent",
        expected_result="??????????????",
        actual_result="????????????????",
        failed_step_no=4,
        failed_step_name="??????",
        artifact_ids=["ART_001", "ART_002", "ART_003", "ART_004"],
    )
    bug_link = BugCaseLink(
        bug_id="BUG-0001",
        run_id=run_context.run_id,
        case_id="LOGIN_002",
        case_result_id="CR_001",
        failed_step_no=4,
        failed_step_name="??????",
        is_primary_case=True,
    )

    generate_output_artifacts(
        config=config,
        run_context=run_context,
        run_info=run_info,
        case_results=[case_result],
        step_results=[step_result],
        artifacts=artifacts,
        bugs=[bug],
        bug_case_links=[bug_link],
    )

    doc_path = run_context.ai_bugs_dir / "BUG-0001.md"
    content = doc_path.read_text(encoding="utf-8")

    assert "## 2. \u95ee\u9898\u5224\u5b9a" in content
    assert "## 8. \u4efb\u52a1\u9886\u53d6\u8bf4\u660e" in content
    assert "## 9. \u4fee\u6539\u8fb9\u754c" in content
    assert "change_scope:" in content
    assert "do_not_change:" in content
    assert "done_definition_1:" in content
    assert "contract_reference: docs/references/ai-bug-handoff/output-contract.md" in content
    assert "- console_log:" in content
    assert "- network_log:" in content

    report_html = (run_context.root_dir / "test_report.html").read_text(encoding="utf-8")

    assert "\u6d4b\u8bd5\u62a5\u544a / Test Report" in report_html
    assert "\u603b\u7528\u4f8b / Total Cases" in report_html
    assert "\u6a21\u5757\u7edf\u8ba1 / Module Summary" in report_html
    assert "\u4f18\u5148\u7ea7\u7edf\u8ba1 / Priority Summary" in report_html
    assert "\u5931\u8d25\u5206\u7c7b\u7edf\u8ba1 / Failure Category Summary" in report_html
    assert "\u8bc1\u636e\u7edf\u8ba1 / Artifact Summary" in report_html
    assert "\u5f02\u5e38\u7528\u4f8b\u6458\u8981 / Failed & Blocked Summary" in report_html
    assert "\u7528\u4f8b\u660e\u7ec6 / Case Details" in report_html
    assert "\u7f3a\u9677\u6458\u8981 / Bug Summary" in report_html
    assert "\u7528\u4f8bID / case_id" in report_html
    assert "\u7f3a\u9677ID / bug_id" in report_html

    case_results_wb = load_workbook(run_context.root_dir / "case_results.xlsx")
    bug_list_wb = load_workbook(run_context.root_dir / "bug_list.xlsx")

    assert "Case_Results - \u7528\u4f8b\u7ed3\u679c" in case_results_wb.sheetnames
    assert "Step_Results - \u6b65\u9aa4\u7ed3\u679c" in case_results_wb.sheetnames
    assert "Case_Stats - \u7528\u4f8b\u7edf\u8ba1" in case_results_wb.sheetnames
    assert case_results_wb["Case_Results - \u7528\u4f8b\u7ed3\u679c"]["A1"].value == "run_id\n\u8fd0\u884cID"
    assert case_results_wb["Step_Results - \u6b65\u9aa4\u7ed3\u679c"]["D1"].value == "step_no\n\u6b65\u9aa4\u5e8f\u53f7"
    assert case_results_wb["Case_Stats - \u7528\u4f8b\u7edf\u8ba1"]["A1"].value == "metric\n\u7edf\u8ba1\u9879"

    case_headers = [cell.value for cell in case_results_wb["Case_Results - \u7528\u4f8b\u7ed3\u679c"][1]]
    step_headers = [cell.value for cell in case_results_wb["Step_Results - \u6b65\u9aa4\u7ed3\u679c"][1]]
    case_stats = {row[0]: row[1] for row in case_results_wb["Case_Stats - \u7528\u4f8b\u7edf\u8ba1"].iter_rows(min_row=2, values_only=True) if row[0]}

    assert "artifact_ids\n\u8bc1\u636eID\u5217\u8868" in case_headers
    assert "attempt_no\n\u5c1d\u8bd5\u6b21\u5e8f" in step_headers
    assert case_stats["case_total"] == 1
    assert case_stats["step_total"] == 1
    assert case_stats["step_failed"] == 1

    assert "Bug_Summary - \u7f3a\u9677\u6c47\u603b" in bug_list_wb.sheetnames
    assert "Bug_Case_Map - \u7f3a\u9677\u7528\u4f8b\u6620\u5c04" in bug_list_wb.sheetnames
    assert "Bug_Evidence - \u7f3a\u9677\u8bc1\u636e" in bug_list_wb.sheetnames
    assert "Bug_Stats - \u7f3a\u9677\u7edf\u8ba1" in bug_list_wb.sheetnames
    assert bug_list_wb["Bug_Summary - \u7f3a\u9677\u6c47\u603b"]["A1"].value == "bug_id\n\u7f3a\u9677ID"
    assert bug_list_wb["Bug_Case_Map - \u7f3a\u9677\u7528\u4f8b\u6620\u5c04"]["B1"].value == "run_id\n\u8fd0\u884cID"
    assert bug_list_wb["Bug_Evidence - \u7f3a\u9677\u8bc1\u636e"].max_row == 5
    assert bug_list_wb["Bug_Stats - \u7f3a\u9677\u7edf\u8ba1"]["A1"].value == "metric\n\u7edf\u8ba1\u9879"

    bug_summary_headers = [cell.value for cell in bug_list_wb["Bug_Summary - \u7f3a\u9677\u6c47\u603b"][1]]
    bug_map_headers = [cell.value for cell in bug_list_wb["Bug_Case_Map - \u7f3a\u9677\u7528\u4f8b\u6620\u5c04"][1]]
    bug_stats = {row[0]: row[1] for row in bug_list_wb["Bug_Stats - \u7f3a\u9677\u7edf\u8ba1"].iter_rows(min_row=2, values_only=True) if row[0]}

    assert "primary_case_id\n\u4e3b\u7528\u4f8bID" in bug_summary_headers
    assert "artifact_count\n\u8bc1\u636e\u6570\u91cf" in bug_summary_headers
    assert "module\n\u6a21\u5757" in bug_map_headers
    assert bug_stats["bug_count"] == 1
    assert bug_stats["artifact_count"] == 4
    assert bug_stats["severity:S2"] == 1
    assert bug_stats["root_cause_category:JS_ERROR"] == 1




def test_test_report_includes_aggregated_statistics_sections(tmp_path: Path) -> None:
    config = AppConfig(
        project_name="DGX Web Auto Test",
        suite_name="report_stats",
        trigger_by="tester",
        base_url="https://example.com",
        output_root=str(tmp_path / "outputs"),
    )
    run_context = prepare_run_context(config)
    summary = ExecutionSummary(
        total_cases=4,
        passed_cases=1,
        failed_cases=1,
        blocked_cases=1,
        skipped_cases=1,
        not_run_cases=0,
        bug_count=0,
    )
    run_info = RunInfo(
        run_id=run_context.run_id,
        project_name=config.project_name,
        suite_name=config.suite_name,
        trigger_by=config.trigger_by,
        start_time="2026-03-23T10:00:00+08:00",
        end_time="2026-03-23T10:01:00+08:00",
        duration_sec=60,
        environment="demo",
        base_url=config.base_url,
        browser=config.browser,
        headless=config.headless,
        execution_summary=summary,
    )
    case_results = [
        CaseResult(
            case_result_id="CR_PASS",
            run_id=run_context.run_id,
            case_id="AUTH_001",
            case_name="Auth pass",
            module="Auth",
            sub_module="Login",
            priority="P2",
            status="PASSED",
            failure_category="",
            expected_summary="Pass",
            actual_summary="Pass",
        ),
        CaseResult(
            case_result_id="CR_FAIL",
            run_id=run_context.run_id,
            case_id="AUTH_002",
            case_name="Auth fail",
            module="Auth",
            sub_module="Login",
            priority="P1",
            status="FAILED",
            failure_category="JS_ERROR",
            expected_summary="Submit succeeds",
            actual_summary="Click crashes",
            failed_step_no=3,
            failed_step_name="Click submit",
            failure_message="TypeError",
            bug_candidate=True,
        ),
        CaseResult(
            case_result_id="CR_BLOCK",
            run_id=run_context.run_id,
            case_id="BILL_001",
            case_name="Billing blocked",
            module="Billing",
            sub_module="Checkout",
            priority="P1",
            status="BLOCKED",
            failure_category="DEPENDENCY_BLOCKED",
            expected_summary="Checkout runs",
            actual_summary="Blocked by dependency",
            failure_message="???????",
        ),
        CaseResult(
            case_result_id="CR_SKIP",
            run_id=run_context.run_id,
            case_id="SEARCH_001",
            case_name="Search skipped",
            module="Search",
            sub_module="List",
            priority="P3",
            status="SKIPPED",
            failure_category="",
            expected_summary="Skip",
            actual_summary="Skip",
        ),
    ]
    artifacts = [
        ArtifactRecord(
            artifact_id="ART_001",
            run_id=run_context.run_id,
            case_id="AUTH_002",
            step_result_id="SR_001",
            artifact_type="screenshot",
            file_path=str(run_context.screenshots_dir / "AUTH_002_failure.png"),
        ),
        ArtifactRecord(
            artifact_id="ART_002",
            run_id=run_context.run_id,
            case_id="AUTH_002",
            step_result_id="SR_001",
            artifact_type="console_log",
            file_path=str(run_context.console_logs_dir / "AUTH_002_failure.log"),
        ),
        ArtifactRecord(
            artifact_id="ART_003",
            run_id=run_context.run_id,
            case_id="BILL_001",
            step_result_id="SR_002",
            artifact_type="screenshot",
            file_path=str(run_context.screenshots_dir / "BILL_001_failure.png"),
        ),
    ]

    generate_output_artifacts(
        config=config,
        run_context=run_context,
        run_info=run_info,
        case_results=case_results,
        step_results=[],
        artifacts=artifacts,
        bugs=[],
        bug_case_links=[],
    )

    report_html = (run_context.root_dir / "test_report.html").read_text(encoding="utf-8")

    assert "\u6a21\u5757\u7edf\u8ba1 / Module Summary" in report_html
    assert "\u4f18\u5148\u7ea7\u7edf\u8ba1 / Priority Summary" in report_html
    assert "\u5931\u8d25\u5206\u7c7b\u7edf\u8ba1 / Failure Category Summary" in report_html
    assert "\u8bc1\u636e\u7edf\u8ba1 / Artifact Summary" in report_html
    assert re.search(r"<tr><td>Auth</td><td>2</td><td>1</td><td>1</td><td>0</td><td>0</td><td>0</td></tr>", report_html)
    assert re.search(r"<tr><td>P1</td><td>2</td><td>0</td><td>1</td><td>1</td><td>0</td><td>0</td></tr>", report_html)
    assert re.search(r"<tr><td>JS_ERROR</td><td>1</td></tr>", report_html)
    assert re.search(r"<tr><td>DEPENDENCY_BLOCKED</td><td>1</td></tr>", report_html)
    assert re.search(r"<tr><td>screenshot</td><td>2</td></tr>", report_html)
    assert re.search(r"<tr><td>console_log</td><td>1</td></tr>", report_html)



def test_excel_outputs_include_richer_fields_and_stats(tmp_path: Path) -> None:
    config = AppConfig(
        project_name="DGX Web Auto Test",
        suite_name="excel_stats",
        trigger_by="tester",
        base_url="https://example.com",
        output_root=str(tmp_path / "outputs"),
    )
    run_context = prepare_run_context(config)
    summary = ExecutionSummary(
        total_cases=2,
        passed_cases=1,
        failed_cases=1,
        blocked_cases=0,
        skipped_cases=0,
        not_run_cases=0,
        bug_count=1,
    )
    run_info = RunInfo(
        run_id=run_context.run_id,
        project_name=config.project_name,
        suite_name=config.suite_name,
        trigger_by=config.trigger_by,
        start_time="2026-03-23T11:00:00+08:00",
        end_time="2026-03-23T11:01:00+08:00",
        duration_sec=60,
        environment="demo",
        base_url=config.base_url,
        browser=config.browser,
        headless=config.headless,
        execution_summary=summary,
    )
    case_results = [
        CaseResult(
            case_result_id="CR_001",
            run_id=run_context.run_id,
            case_id="AUTH_001",
            case_name="Auth fail",
            module="Auth",
            sub_module="Login",
            priority="P1",
            status="FAILED",
            failure_category="ASSERTION_FAILED",
            expected_summary="Login succeeds",
            actual_summary="Button stays disabled",
            failed_step_no=2,
            failed_step_name="Assert login status",
            failure_message="Assertion failed",
            bug_candidate=True,
            dedup_key="assert|auth",
            artifact_ids=["ART_001", "ART_002"],
            suspected_layer="frontend",
            bug_id="BUG-0001",
            retry_count=1,
        ),
        CaseResult(
            case_result_id="CR_002",
            run_id=run_context.run_id,
            case_id="AUTH_002",
            case_name="Auth pass",
            module="Auth",
            sub_module="Login",
            priority="P2",
            status="PASSED",
            failure_category="",
            expected_summary="Pass",
            actual_summary="Pass",
        ),
    ]
    step_results = [
        StepResult(
            step_result_id="SR_001",
            run_id=run_context.run_id,
            case_id="AUTH_001",
            step_no=2,
            step_type="assert",
            step_name="Assert login status",
            status="FAILED",
            assert_key="assert_element_enabled",
            target="login_button",
            page_name="LoginPage",
            expected="enabled",
            actual="disabled",
            error_type="ASSERTION_FAILED",
            error_message="Element is disabled",
            page_url="https://example.com/login",
            locator_used_type="testid",
            locator_used_value="login-button",
            component_hint="Login.Button",
            frontend_file_hint="src/components/Login.tsx",
            api_hint="POST /api/login",
            attempt_no=2,
        )
    ]
    artifacts = [
        ArtifactRecord(
            artifact_id="ART_001",
            run_id=run_context.run_id,
            case_id="AUTH_001",
            step_result_id="SR_001",
            artifact_type="screenshot",
            file_path=str(run_context.screenshots_dir / "AUTH_001_failure.png"),
        ),
        ArtifactRecord(
            artifact_id="ART_002",
            run_id=run_context.run_id,
            case_id="AUTH_001",
            step_result_id="SR_001",
            artifact_type="console_log",
            file_path=str(run_context.console_logs_dir / "AUTH_001_failure.log"),
        ),
    ]
    bug = BugRecord(
        bug_id="BUG-0001",
        run_id=run_context.run_id,
        title="Auth - login assertion failed",
        module="Auth",
        sub_module="Login",
        severity="S2",
        priority="P1",
        status="NEW",
        root_cause_category="ASSERTION_FAILED",
        suspected_layer="frontend",
        affected_case_ids=["AUTH_001"],
        affected_case_count=1,
        dedup_key="assert|auth",
        expected_result="Login succeeds",
        actual_result="Button stays disabled",
        failed_step_no=2,
        failed_step_name="Assert login status",
        artifact_ids=["ART_001", "ART_002"],
    )
    bug_link = BugCaseLink(
        bug_id="BUG-0001",
        run_id=run_context.run_id,
        case_id="AUTH_001",
        case_result_id="CR_001",
        failed_step_no=2,
        failed_step_name="Assert login status",
        is_primary_case=True,
    )

    generate_output_artifacts(
        config=config,
        run_context=run_context,
        run_info=run_info,
        case_results=case_results,
        step_results=step_results,
        artifacts=artifacts,
        bugs=[bug],
        bug_case_links=[bug_link],
    )

    case_results_wb = load_workbook(run_context.root_dir / "case_results.xlsx")
    bug_list_wb = load_workbook(run_context.root_dir / "bug_list.xlsx")

    case_sheet = case_results_wb["Case_Results - 用例结果"]
    step_sheet = case_results_wb["Step_Results - 步骤结果"]
    case_stats_sheet = case_results_wb["Case_Stats - 用例统计"]
    bug_summary_sheet = bug_list_wb["Bug_Summary - 缺陷汇总"]
    bug_map_sheet = bug_list_wb["Bug_Case_Map - 缺陷用例映射"]
    bug_stats_sheet = bug_list_wb["Bug_Stats - 缺陷统计"]

    case_row = _sheet_row_as_dict(case_sheet)
    step_row = _sheet_row_as_dict(step_sheet)
    bug_summary_row = _sheet_row_as_dict(bug_summary_sheet)
    bug_map_row = _sheet_row_as_dict(bug_map_sheet)
    case_stats = {row[0]: row[1] for row in case_stats_sheet.iter_rows(min_row=2, values_only=True) if row[0]}
    bug_stats = {row[0]: row[1] for row in bug_stats_sheet.iter_rows(min_row=2, values_only=True) if row[0]}

    assert case_row["artifact_ids"] == "ART_001,ART_002"
    assert step_row["step_result_id"] == "SR_001"
    assert step_row["attempt_no"] == 2
    assert case_stats["case_failed"] == 1
    assert case_stats["case_passed"] == 1
    assert case_stats["max_retry_count"] == 1
    assert bug_summary_row["primary_case_id"] == "AUTH_001"
    assert bug_summary_row["artifact_count"] == 2
    assert bug_map_row["module"] == "Auth"
    assert bug_map_row["case_status"] == "FAILED"
    assert bug_stats["artifact_count"] == 2
    assert bug_stats["root_cause_category:ASSERTION_FAILED"] == 1
